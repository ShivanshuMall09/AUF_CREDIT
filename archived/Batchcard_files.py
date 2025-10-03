# from prettytable import PrettyTable
# from openpyxl.utils import get_column_letter
# import shutil
# import os
# import time
# import glob
# from openpyxl.styles import PatternFill, Border, Side, Alignment
# from openpyxl.utils.dataframe import dataframe_to_rows
# import win32com.client as win32
# import random
# import datetime
# import hashlib
# from pathlib import Path
# from datetime import date
# import pandas as pd
# from babel.numbers import format_currency
# import traceback
# from openpyxl import load_workbook
# import importlib.machinery
# import importlib.util
# import warnings

# # Suppress FutureWarnings
# warnings.simplefilter(action='ignore', category=FutureWarning)

# def material_issueance_file(file_to_process):
#             try:
#                 import re
#                 input_dir_path = os.getcwd()
#                 input_file_path = os.path.join(input_dir_path, file_to_process)
#                 # Extract date and count from the file name (e.g., "03.01.2025_352")
#                 match = re.search(r"(\d{2}\-\d{2}\-\d{4})_(\d+)", file_to_process)
#                 if match:
#                     date_part = match.group(1)  # Extract date (e.g., "03.01.2025")
#                     count_part = match.group(2)  # Extract count (e.g., "352")
#                     output_file_name = f"AUC_Material_Issuance_Detail_{date_part}--{count_part}.xlsx"
#                 else:
#                     print("Date and count could not be extracted from the file name.")
#                     return

#                 # Read the input Excel file
#                 input_df = pd.read_excel(input_file_path)

#                 # Group by "Card Body A/W Ref.No." (or "ARTWORK NO.") and sum the "Quantity"
#                 grouped_df = input_df.groupby(["ARTWORK NO.", "BIN"], as_index=False).agg({
#                     "QTY": "sum"  # Sum the "Quantity"
#                 })
#                 # print(grouped_df)
#                 # Define output headers
#                 output_headers = [
#                     "Item Requested",                                                                                                
#                     "Card Body A/W Ref.No.",
#                     "BIN",
#                     "Req Quantity",
#                     "Issued Quantity",
#                     "Request By Name",
#                     "Request By Sign",
#                     "Date & Time",
#                     "Issued / Name",
#                     "Issued / Sign",
#                     "Date & Time1",
#                     "Remark If any."
#                 ]

#                 # Create the output DataFrame with the required headers
#                 output_df = pd.DataFrame(columns=output_headers)

#                 # Populate the output DataFrame
#                 output_df["Item Requested"] = ["Card"] * len(grouped_df)  # Fixed value "Card"
#                 output_df["Card Body A/W Ref.No."] = grouped_df["ARTWORK NO."]  # Map from grouped data
#                 output_df["BIN"] = grouped_df["BIN"]  # Map from grouped data
#                 output_df["Req Quantity"] = grouped_df["QTY"]  # Map summed "Quantity"

#                 # Add empty columns for other headers
#                 output_df["Issued Quantity"] = ''
#                 output_df["Request By Name"] = ''
#                 output_df["Request By Sign"] = ''
#                 output_df["Date & Time"] = ''
#                 output_df["Issued / Name"] = ''
#                 output_df["Issued / Sign"] = ''
#                 output_df["Date & Time1"] = ''
#                 output_df["Remark If any."] = ''

#                 # Compute the sum of the "Quantity" column
#                 total_quantity = grouped_df["QTY"].sum()

#                 # Save the output file in the same directory as the input file
#                 output_file_path = os.path.join(input_dir_path, output_file_name)
#                 output_df.to_excel(output_file_path, index=False)

#                 # Open the generated Excel file and apply formatting
#                 wb = load_workbook(output_file_path)
#                 ws = wb.active

#                 # Set sheet view to 85%
#                 ws.sheet_view.zoomScale = 85

#                 # Make the first row gray
#                 gray_fill = PatternFill(start_color="BEBEBE", end_color="BEBEBE", fill_type="solid")
#                 for cell in ws[1]:
#                     cell.fill = gray_fill

#                 # Freeze the first row
#                 ws.freeze_panes = "A2"

#                 # Add borders to all data cells
#                 border = Border(left=Side(border_style="thin"),
#                                 right=Side(border_style="thin"),
#                                 top=Side(border_style="thin"),
#                                 bottom=Side(border_style="thin"))

#                 for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
#                     for cell in row:
#                         cell.border = border
#                         # Center align all cells
#                         cell.alignment = Alignment(horizontal="center", vertical="center")

#                 # Auto-fit column width based on the maximum length of data in each column
#                 for col in ws.columns:
#                     max_length = 0
#                     column = col[0].column_letter  # Get column name (e.g., 'A', 'B', etc.)
#                     for cell in col:
#                         try:
#                             # For both string and numeric values, get the max length
#                             if cell.value:
#                                 max_length = max(max_length, len(str(cell.value)))
#                         except:
#                             pass
#                     # Adjust width with extra padding for readability
#                     adjusted_width = (max_length + 2)
#                     ws.column_dimensions[column].width = adjusted_width

#                 # Save the formatted output
#                 wb.save(output_file_path)

#                 print(f"\nOutput file saved at: {output_file_path}")
#                 print(f"\nTotal Data Count : {total_quantity}")

#             except Exception as e:
#                 print(f"An error occurred: {e}")


# def reconcillation_file(file_to_process):
#             try:
#                 import re
#                 input_dir_path = os.getcwd()
#                 input_file_path = os.path.join(input_dir_path, file_to_process)
#                 # Extract date and count from the file name (e.g., "03.01.2025_352")
#                 match = re.search(r"(\d{2}\-\d{2}\-\d{4})_(\d+)", file_to_process)
#                 if match:
#                     date_part = match.group(1)  # Extract date (e.g., "03.01.2025")
#                     count_part = match.group(2)  # Extract count (e.g., "352")
#                     output_file_name = f"AUC_Reconcillation_Detail_{date_part}--{count_part}.xlsx"
#                 else:
#                     print("Date and count could not be extracted from the file name.")
#                     return

#                 # Read the input Excel file
#                 input_df = pd.read_excel(input_file_path)

#                 # Group by "Card Body A/W Ref.No." (or "ARTWORK NO.") and sum the "Quantity"
#                 grouped_df = input_df.groupby(["JOB_SETUP_NAME","ARTWORK NO.", "BIN"], as_index=False).agg({
#                     "QTY": "sum"  # Sum the "Quantity"
#                 })
#                 # print(grouped_df)
#                 # Define output headers
#                 output_headers = [
#                     "Card Setup File Name",                                                                                                
#                     "Card Body A/W Ref.No.",
#                     "BIN",
#                     "Issued cards Qty.",
#                     "Total Personalized Cards",
#                     "Reject Card Qty.",
#                     "Left Over",
#                     "Sample",
#                     "Handover Given By name & sign",
#                     "Handover Taken By name & sign",
#                     "Date & Time",
#                     "Remark If any."
#                 ]

#                 # Create the output DataFrame with the required headers
#                 output_df = pd.DataFrame(columns=output_headers)

#                 # Populate the output DataFrame
#                 output_df["Card Setup File Name"] = grouped_df["JOB_SETUP_NAME"]   # Fixed value "Card"
#                 output_df["Card Body A/W Ref.No."] = grouped_df["ARTWORK NO."]  # Map from grouped data
#                 output_df["BIN"] = grouped_df["BIN"]  # Map from grouped data
#                 output_df["Issued cards Qty."] = grouped_df["QTY"]  # Map summed "Quantity"

#                 # Add empty columns for other headers
                
#                 output_df["Total Personalized Cards"] = ''
#                 output_df["Reject Card Qty."] = ''
#                 output_df["Left Over"] = ''
#                 output_df["Sample"] = ''
#                 output_df["Handover Given By name & sign"] = ''
#                 output_df["Handover Taken By name & sign"] = ''
#                 output_df["Date & Time"] = ''
#                 output_df["Remark If any."] = ''

#                 # Compute the sum of the "Quantity" column
#                 total_quantity = grouped_df["QTY"].sum()

#                 # Save the output file in the same directory as the input file
#                 output_file_path = os.path.join(input_dir_path, output_file_name)
#                 output_df.to_excel(output_file_path, index=False)

#                 # Open the generated Excel file and apply formatting
#                 wb = load_workbook(output_file_path)
#                 ws = wb.active

#                 # Set sheet view to 85%
#                 ws.sheet_view.zoomScale = 85

#                 # Make the first row gray
#                 gray_fill = PatternFill(start_color="BEBEBE", end_color="BEBEBE", fill_type="solid")
#                 for cell in ws[1]:
#                     cell.fill = gray_fill

#                 # Freeze the first row
#                 ws.freeze_panes = "A2"

#                 # Add borders to all data cells
#                 border = Border(left=Side(border_style="thin"),
#                                 right=Side(border_style="thin"),
#                                 top=Side(border_style="thin"),
#                                 bottom=Side(border_style="thin"))

#                 for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
#                     for cell in row:
#                         cell.border = border
#                         # Center align all cells
#                         cell.alignment = Alignment(horizontal="center", vertical="center")

#                 # Auto-fit column width based on the maximum length of data in each column
#                 for col in ws.columns:
#                     max_length = 0
#                     column = col[0].column_letter  # Get column name (e.g., 'A', 'B', etc.)
#                     for cell in col:
#                         try:
#                             # For both string and numeric values, get the max length
#                             if cell.value:
#                                 max_length = max(max_length, len(str(cell.value)))
#                         except:
#                             pass
#                     # Adjust width with extra padding for readability
#                     adjusted_width = (max_length + 2)
#                     ws.column_dimensions[column].width = adjusted_width

#                 # Save the formatted output
#                 wb.save(output_file_path)

#                 print(f"\nOutput file saved at: {output_file_path}")
#                 print(f"\nTotal Data Count : {total_quantity}")

#             except Exception as e:
#                 print(f"An error occurred: {e}")

# def material_Return_file(file_to_process):
#             try:
#                 import re
#                 input_dir_path = os.getcwd()
#                 input_file_path = os.path.join(input_dir_path, file_to_process)
#                 # Extract date and count from the file name (e.g., "03.01.2025_352")
#                 match = re.search(r"(\d{2}\-\d{2}\-\d{4})_(\d+)", file_to_process)
#                 if match:
#                     date_part = match.group(1)  # Extract date (e.g., "03.01.2025")
#                     count_part = match.group(2)  # Extract count (e.g., "352")
#                     output_file_name = f"AUC_Material_Return_Detail_{date_part}--{count_part}.xlsx"
#                 else:
#                     print("Date and count could not be extracted from the file name.")
#                     return

#                 # Read the input Excel file
#                 input_df = pd.read_excel(input_file_path)

#                 # Group by "Card Body A/W Ref.No." (or "ARTWORK NO.") and sum the "Quantity"
#                 grouped_df = input_df.groupby(["ARTWORK NO.", "BIN"], as_index=False).agg({
#                     "QTY": "sum"  # Sum the "Quantity"
#                 })
#                 # print(grouped_df)
#                 # Define output headers
#                 output_headers = [
#                     "Item Requested",                                                                                                
#                     "Card Body A/W Ref.No.",
#                     "BIN",
#                     "Returned Quantity",
#                     "Receive Quantity",
#                     "Return / Name",
#                     "Return / Sign",
#                     "Date & Time",
#                     "Receive / Name",
#                     "Receive / Sign",
#                     "Date & Time",
#                     "Remark If any."
#                 ]

#                 # Create the output DataFrame with the required headers
#                 output_df = pd.DataFrame(columns=output_headers)

#                 # Populate the output DataFrame
#                 output_df["Item Requested"] = ["Card"] * len(grouped_df)  # Fixed value "Card"
#                 output_df["Card Body A/W Ref.No."] = grouped_df["ARTWORK NO."]  # Map from grouped data
#                 output_df["BIN"] = grouped_df["BIN"]  # Map from grouped data
#                 output_df["Returned Quantity"] = ''  # Map summed "Quantity"

#                 # Add empty columns for other headers
#                 output_df["Receive Quantity"] = ''
#                 output_df["Return / Name"] = ''
#                 output_df["Return / Sign"] = ''
#                 output_df["Date & Time"] = ''
#                 output_df["Receive / Name"] = ''
#                 output_df["Receive / Sign"] = ''
#                 output_df["Date & Time"] = ''
#                 output_df["Remark If any."] = ''

#                 # Compute the sum of the "Quantity" column
#                 total_quantity = grouped_df["QTY"].sum()

#                 # Save the output file in the same directory as the input file
#                 output_file_path = os.path.join(input_dir_path, output_file_name)
#                 output_df.to_excel(output_file_path, index=False)

#                 # Open the generated Excel file and apply formatting
#                 wb = load_workbook(output_file_path)
#                 ws = wb.active

#                 # Set sheet view to 85%
#                 ws.sheet_view.zoomScale = 85

#                 # Make the first row gray
#                 gray_fill = PatternFill(start_color="BEBEBE", end_color="BEBEBE", fill_type="solid")
#                 for cell in ws[1]:
#                     cell.fill = gray_fill

#                 # Freeze the first row
#                 ws.freeze_panes = "A2"

#                 # Add borders to all data cells
#                 border = Border(left=Side(border_style="thin"),
#                                 right=Side(border_style="thin"),
#                                 top=Side(border_style="thin"),
#                                 bottom=Side(border_style="thin"))

#                 for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
#                     for cell in row:
#                         cell.border = border
#                         # Center align all cells
#                         cell.alignment = Alignment(horizontal="center", vertical="center")

#                 # Auto-fit column width based on the maximum length of data in each column
#                 for col in ws.columns:
#                     max_length = 0
#                     column = col[0].column_letter  # Get column name (e.g., 'A', 'B', etc.)
#                     for cell in col:
#                         try:
#                             # For both string and numeric values, get the max length
#                             if cell.value:
#                                 max_length = max(max_length, len(str(cell.value)))
#                         except:
#                             pass
#                     # Adjust width with extra padding for readability
#                     adjusted_width = (max_length + 2)
#                     ws.column_dimensions[column].width = adjusted_width

#                 # Save the formatted output
#                 wb.save(output_file_path)

#                 print(f"\nOutput file saved at: {output_file_path}")
#                 print(f"\nTotal Data Count : {total_quantity}")

#             except Exception as e:
#                 print(f"An error occurred: {e}")







# input_file = glob.glob('AUF_CREDIT_BATCHCARD_*.xlsx')
# material_issueance_file(input_file[0])
# reconcillation_file(input_file[0])
# material_Return_file(input_file[0])




import os
import re
import glob
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

def apply_formatting(ws):
    # Zoom
    ws.sheet_view.zoomScale = 85
    # Freeze first row
    ws.freeze_panes = "A2"
    # Gray header fill
    gray_fill = PatternFill(start_color="BEBEBE", end_color="BEBEBE", fill_type="solid")
    for cell in ws[1]:
        cell.fill = gray_fill
    # Border and center alignment
    border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")
    # Auto-fit columns
    for col in ws.columns:
        max_length = max((len(str(cell.value)) if cell.value else 0) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_length + 2

def process_combined_file(file_to_process):
    try:
        input_path = os.path.join(os.getcwd(), file_to_process)
        match = re.search(r"(\d{2}\-\d{2}\-\d{4})_(\d+)", file_to_process)
        if not match:
            print("Date and count could not be extracted from the file name.")
            return
        date_part, count_part = match.group(1), match.group(2)
        output_file_name = f"AUC_Material_Report_{date_part}--{count_part}.xlsx"

        # Read input file
        df = pd.read_excel(input_path)

        # Create Excel workbook
        wb = Workbook()
        del wb["Sheet"]  

        # --------------------------
        # Sheet 1: Material Issuance
        # --------------------------
        grouped_issuance = df.groupby(["ARTWORK NO.", "BIN"], as_index=False)["QTY"].sum()
        issuance_df = pd.DataFrame({
            "Item Requested": "Card",
            "Card Body A/W Ref.No.": grouped_issuance["ARTWORK NO."],
            "BIN": grouped_issuance["BIN"],
            "Req Quantity": grouped_issuance["QTY"],
            "Issued Quantity": "",
            "Request By Name": "",
            "Request By Sign": "",
            "Date & Time": "",
            "Issued / Name": "",
            "Issued / Sign": "",
            "Date & Time1": "",
            "Remark If any.": ""
        })

        ws1 = wb.create_sheet("Material Issuance")
        for r in dataframe_to_rows(issuance_df, index=False, header=True):
            ws1.append(r)
        apply_formatting(ws1)

        # --------------------------
        # Sheet 2: Reconciliation
        # --------------------------
        grouped_recon = df.groupby(["JOB_SETUP_NAME", "ARTWORK NO.", "BIN"], as_index=False)["QTY"].sum()
        recon_df = pd.DataFrame({
            "Card Setup File Name": grouped_recon["JOB_SETUP_NAME"],
            "Card Body A/W Ref.No.": grouped_recon["ARTWORK NO."],
            "BIN": grouped_recon["BIN"],
            "Issued cards Qty.": grouped_recon["QTY"],
            "Total Personalized Cards": "",
            "Reject Card Qty.": "",
            "Left Over": "",
            "Sample": "",
            "Handover Given By name & sign": "",
            "Handover Taken By name & sign": "",
            "Date & Time": "",
            "Remark If any.": ""
        })

        ws2 = wb.create_sheet("Reconciliation")
        for r in dataframe_to_rows(recon_df, index=False, header=True):
            ws2.append(r)
        apply_formatting(ws2)

        # --------------------------
        # Sheet 3: Material Return
        # --------------------------
        grouped_return = df.groupby(["ARTWORK NO.", "BIN"], as_index=False)["QTY"].sum()
        return_df = pd.DataFrame({
            "Item Requested": "Card",
            "Card Body A/W Ref.No.": grouped_return["ARTWORK NO."],
            "BIN": grouped_return["BIN"],
            "Returned Quantity": "",
            "Receive Quantity": "",
            "Return / Name": "",
            "Return / Sign": "",
            "Date & Time": "",
            "Receive / Name": "",
            "Receive / Sign": "",
            "Date & Time": "",
            "Remark If any.": ""
        })

        ws3 = wb.create_sheet("Material Return")
        for r in dataframe_to_rows(return_df, index=False, header=True):
            ws3.append(r)
        apply_formatting(ws3)

        # Save file
        wb.save(output_file_name)
        print(f"\n✅ All sheets saved to one file: {output_file_name}")
        # print(f"📦 Total Material Issuance Qty: {grouped_issuance['QTY'].sum()}")
        # print(f"📦 Total Reconciliation Qty: {grouped_recon['QTY'].sum()}")
        # print(f"📦 Total Material Return Qty: {grouped_return['QTY'].sum()}")

    except Exception as e:
        print(f"❌ Error: {e}")

# Run the script
input_files = glob.glob("AUF_CREDIT_BATCHCARD_*.xlsx")
if input_files:
    process_combined_file(input_files[0])
else:
    print("❌ No matching input file found.")
