# -*- coding: utf-8 -*-
"""
Created on Wed April 17 17:24:59 2024

@author: shivanshu.mall
AU_Credit_IXIGO_PP_NEW_ISSUE_V1.0.7.py:17-07-2024routing_code changed to  '921-413' from  '921-417'
AU_Credit_IXIGO_PP_NEW_ISSUE_V1.0.6.py:17-07-2024[Changing the RTO addres in DElhivery courier connection file request it raised by aditya on mail and also adding the new issue in the file name]
AU_Credit_IXIGO_PP_NEW_ISSUE_V1.0.5.py:06-06-2024[Fixing the complemantary lounge access on the basis of the logo]
AU_Credit_IXIGO_PP_NEW_ISSUE_V1.0.4.py:29-05-2024[Adding acount_holder name in the mis file]
AU_Credit_IXIGO_PP_NEW_ISSUE_V1.0.3.py:15-05-2024[Adding the account holder name in the ff file ]
AU_Credit_IXIGO_PP_NEW_ISSUE_V1.0.2.py:09-05-2024[Seprating the ff and  embo on the product varient]
AU_Credit_IXIGO_PP_NEW_ISSUE_V1.0.1.py:23-04-2024[Developing the tool on the basis of given requirement]

"""

import os
import glob
from openpyxl import load_workbook
import traceback 
import pandas as pd
import sys
from pathlib import Path
from datetime import datetime, timedelta

try:
    print("DATA PROCESSING STARTS.....")
    cwd = os.getcwd()
    #del_file1 = glob.glob("*.txt*")
    cwd = os.getcwd()
    
    header = 0
    
    pp_ixigo_count=[]
    pp_zenith_count=[]
    pp_VETTA_count=[]
    ts = datetime.now()
    
    date = ts.strftime("%d-%m-%Y")
    ptime = ts.strftime("%d.%m.%Y_%H%M%S")
    print(ptime)
    x = (ts)
    pp_num =[]
    with open('config/AUC_batch_series', 'r') as fin:
        batch_series_data = fin.read().splitlines(True)
        batch_series = batch_series_data[0:1]
    fin.close()
    with open('config/AUC_batch_series', 'w') as fout:
        fout.writelines(batch_series_data[1:])
    fout.close()
    listToStr = ' '.join([str(elem) for elem in batch_series])
    batch_number = listToStr[0:5]
    
    with open('config/IXIGO_PP_NUMBER.csv', 'r') as IXIGOin, open('config/IXIGO_PP_NUMBER_USING.csv', 'w') as IXIGOout:
        contents = IXIGOin.readlines()
        ixigo_Total_count = len(contents)
        IXIGOout.writelines(contents)
    IXIGOin.close()
    IXIGOout.close()
    
    
    with open('config/ZENITH_PP_NUMBER.csv', 'r') as zenithin, open('config/ZENITH_PP_NUMBER_USING.csv', 'w') as zenithout:
        contents = zenithin.readlines()
        ZENITH_Total_count = len(contents)
        zenithout.writelines(contents)
    zenithin.close()
    zenithout.close()
    
    with open('config/VETTA_PP_NUMBER.csv', 'r') as vettain, open('config/VETTA_PP_NUMBER_USING.csv', 'w') as vettaout:
        contents = vettain.readlines()
        VETTA_Total_count = len(contents)
        vettaout.writelines(contents)
    vettain.close()
    vettaout.close()
    
    def convert_to_read_only(file_path):
        try:
            workbook = load_workbook(file_path)
            for sheet in workbook.sheetnames:
                worksheet = workbook[sheet]
                worksheet.protection.sheet = True
            workbook.save(file_path)
            workbook.close()
        except Exception as e:
            print(f"Error converting '{file_path}': {e}")
    
    def pp_dup_check(number):
        pp_num = []  
        with open('config/AUF_CREDIT_PP_CONSOLE.txt', 'r') as file_read:
            contents = file_read.readlines()
            new_contents = [x.strip() for x in contents]  # Remove newline characters
            if number in new_contents:
                print("PP duplicate number found: " + number)
            elif number in pp_num:
                print("PP duplicate number found: " + number)
            else:
                # print("PP number not found in the file: " + number)
                pp_num.append(number)
                with open('config/AUF_CREDIT_PP_CONSOLE.txt','a') as file_write:
                    for line in pp_num:
                        file_write.write(line+'\n')
                file_write.close()
        return number
    
    
    
    
    def pp_num(bin_check):
        if bin_check == '465523' :#Vetta cards
            with open('config/VETTA_PP_NUMBER_USING.csv', 'r') as vetta_pp_in:
                vetta_pp_total    = vetta_pp_in.readline()
                pp_cardnumber_acc = vetta_pp_total.rstrip()
                # print(pp_cardnumber_acc + 'vetta pp')
                vetta_pp_unused   = vetta_pp_in.readlines()
            vetta_pp_in.close()    
               
            with open('config/VETTA_PP_NUMBER_USING.csv', 'w') as vetta_pp_out:
                vetta_pp_out.writelines(vetta_pp_unused)
                pp_VETTA_count.append(str(vetta_pp_total))
                
            vetta_pp_out.close()
            
        elif bin_check == '457036' :#Zenith cards
            with open('config/ZENITH_PP_NUMBER_USING.csv', 'r') as zenith_pp_in:
                zenith_pp_total   = zenith_pp_in.readline()
                pp_cardnumber_acc = zenith_pp_total.rstrip()
                # print(pp_cardnumber_acc + 'zenith pp')
                zenith_pp_unused  = zenith_pp_in.readlines()
            zenith_pp_in.close()    
            with open('config/ZENITH_PP_NUMBER_USING.csv', 'w') as zenith_pp_out:
                zenith_pp_out.writelines(zenith_pp_unused)
                pp_zenith_count.append(str(zenith_pp_total))
            zenith_pp_out.close()
        #print(pp_cardnumber_acc)
        # return pp_cardnumber_acc
    
        elif bin_check == '406977' :#IXIGO
            with open('config/IXIGO_PP_NUMBER_USING.csv', 'r') as IXIGO_pp_in:
                IXIGO_pp_total   = IXIGO_pp_in.readline()
                pp_cardnumber_acc = IXIGO_pp_total.rstrip()
                # print(pp_cardnumber_acc + 'zenith pp')
                IXIGO_pp_unused  = IXIGO_pp_in.readlines()
            IXIGO_pp_in.close()    
            with open('config/IXIGO_PP_NUMBER_USING.csv', 'w') as IXIGO_pp_out:
                IXIGO_pp_out.writelines(IXIGO_pp_unused)
                pp_ixigo_count.append(str(IXIGO_pp_total))
            IXIGO_pp_out.close()
        #print(pp_cardnumber_acc)
        return pp_cardnumber_acc
    
    def awb_assign(cust_courier):
        awb_number = ''
        #print(cust_courier)
        if cust_courier.upper() == 'BLUEDART':
            #print("yesy")
            with open('config/BD_AWB.txt', 'r') as bdfin:
                bdawb = bdfin.readline()
                awb_number=bdawb.rstrip()
                bdawb1 = bdfin.readlines()
                bdremainingnew = len(bdawb1)
                if bdremainingnew < 1:
                    with open ("Error.txt",'w') as error_file:
                        error_file.write("!!!!BD AWB NUMBER FINISHED!!!!!!")
                    error_file
                    
                    sys.exit("!!!!BD AWB NUMBER FINISHED!!!!!!")
                    
            with open('config/BD_AWB.txt', 'w') as bdfout:
                 bdfout.writelines(bdawb1)
                 
                        
        elif cust_courier == 'DELHIVERY':
            with open('config/DL_AWB.txt', 'r') as dlfin:
                dlawb = dlfin.readline()
                awb_number=dlawb.rstrip()
                dlawb_1 = dlfin.readlines()
                awbremainingnew = len(dlawb_1)
                # print(awbremainingnew)
                if awbremainingnew < 1:
                    with open ("Error.txt",'w') as error_file:
                        error_file.write("!!!!DELHIVERY AWB NUMBER FINISHED!!!!!!")
                    error_file
                    sys.exit("!!!!DELHIVERY AWB NUMBER FINISHED!!!!!!")
            dlawb_2 = list(map(str, dlawb_1))
            dlawb_2 = ''.join(dlawb_2)        
            with open('config/DL_AWB.txt', 'w') as dlfout:
                dlfout.write(dlawb_2)
                 
                
        elif cust_courier == 'Speedpost':
            with open('config/IP_AWB.txt', 'r') as ipfin:
                # print("speed")
                ipawb = ipfin.readline()
                awb_number = ipawb.rstrip()
                ipawb_1 = ipfin.readlines()
                ipawbremainingnew = len(ipawb_1)
                # print(ipawbremainingnew)
                if ipawbremainingnew < 1:
                    with open ("Error.txt",'w') as error_file:
                        error_file.write("!!!!IP AWB NUMBER FINISHED!!!!!!")
                    error_file
                    sys.exit("!!!!IP AWB NUMBER FINISHED!!!!!!")
            ipawb_2 = list(map(str, ipawb_1))
            ipawb_2 = ''.join(ipawb_2)
            
            with open('config/IP_AWB.txt', 'w') as ipfout:
                ipfout.write(ipawb_2)
        # print(awb_number)
        return(awb_number)
    
    def courier_mode(cust_pincode):
        
        with open('config/PINCODE_MASTER.csv', 'r') as pin:
            routing_codewe = pin.readlines()
            flag = 0
            routing_code = ''
            for line in routing_codewe:
                pincodedata = line.split(',')
                if cust_pincode != '':
                    # and pincodedata[4]=='BLUEDART':
                    if cust_pincode in pincodedata[0]:
                        flag = 1
                        courier = pincodedata[2].rstrip()
                        routing_code = pincodedata[1].rstrip()
                        break
            if flag == 0:
                courier = 'Speedpost'
                routing_code = '921-413'
                courier = str(courier)
        pin.close()
        return(courier,routing_code)
    
    def my_sort_embo(line):
        line = line.strip()
        couriersort = line[-2:]
        return couriersort
    
    
    
    def my_sort_ff(line):
        line_fields = line.strip().split('|')
        couriersort = (line_fields[13])
        return couriersort
    
    
    def excel_convertor(name_ff):
        # if name_ff.__contains__('FF'):
        if Path(name_ff).is_file():
            with open(name_ff,'r') as ff_file ,open(name_ff[:-4]+ptime+'.txt','w') as ff_file_out:
                ff_file_out.write('Sr. No|Account Number|Primary Card number|Customer Name|Add-on Card 1|Add-on Card 2|Add-on Card 3|Add-on Card 4|PP  Card No_Primary|PP  Card No_Addon1|PP  Card No_Addon2|PP  Card No_Addon3|Ref no.|AWB. No.|Courier|Courier Code|Card Action|Primary Count|Add-on Count|Total Cards|Bin|Logo|Gender Code|Varient|Address Line 1|Address Line 2|Address Line 3|Address Line 4|City|State Code|Postal Code|Mobile Number|Cust Unique ID|Credit Limit|Statement Date|Extention|AccountHolder_Name|Product Varient|Complimentary Lounge Access|Expiry Date\n')
                contents  = ff_file.readlines()
                # sorting using our custom logic
                contents.sort(key=my_sort_ff)
                sr_no = 1
                for line in contents:
                    s_no = "%04d" % sr_no
                    ff_file_out.write(str(s_no)+'|'+line)
                    sr_no+=1
            ff_file.close()
            ff_file_out.close()
            sr_no-=1
            df = pd.read_csv(name_ff[:-4]+ptime+'.txt',encoding= 'unicode_escape', sep='|', dtype=object)
            df.to_excel(name_ff[:-4]+str(ptime)+'_'+str(sr_no)+'.xlsx', 'Sheet1', index=False)
            os.remove(name_ff)
            os.remove(name_ff[:-4]+ptime+'.txt')
        
        # else:
        #     if Path(name_ff).is_file():
        #         with open(name_ff,'r') as ff_file ,open(name_ff[:-4]+ptime+'.txt','w') as ff_file_out:
        #             ff_file_out.write('Sr. No|Account Number|Primary Card number|Customer Name|Add-on Card 1|Add-on Card 2|Add-on Card 3|Add-on Card 4|PP  Card No_Primary|PP  Card No_Addon1|PP  Card No_Addon2|PP  Card No_Addon3|Ref no.|AWB. No.|Courier|Courier Code|Card Action|Primary Count|Add-on Count|Total Cards|Bin|Logo|Gender Code|Varient|Address Line 1|Address Line 2|Address Line 3|Address Line 4|City|State Code|Postal Code|Mobile Number|Cust Unique ID|Credit Limit|Statement Date|Extention|Product Varient|Complimentary Lounge Access|Expiry Date\n')
        #             contents  = ff_file.readlines()
        #             # sorting using our custom logic
        #             contents.sort(key=my_sort_ff)
        #             sr_no = 1
        #             for line in contents:
        #                 s_no = "%04d" % sr_no
        #                 ff_file_out.write(str(s_no)+'|'+line)
        #                 sr_no+=1
        #         ff_file.close()
        #         ff_file_out.close()
        #         sr_no-=1
        #         df = pd.read_csv(name_ff[:-4]+ptime+'.txt',encoding= 'unicode_escape', sep='|', dtype=object)
        #         df.to_excel(name_ff[:-4]+str(ptime)+'_'+str(sr_no)+'.xlsx', 'Sheet1', index=False)
        #         os.remove(name_ff)
        #         os.remove(name_ff[:-4]+ptime+'.txt')
        
        
    def embo_sorting(name_embo):
        #ext='.txt'
        if Path(name_embo).is_file():
            with open(name_embo,'r') as embo_file_in ,open(name_embo[2:],'w') as embo_file_out:
                contents  = embo_file_in.readlines()
                # sorting using our custom logic
                contents.sort(key=my_sort_embo)
                for line in contents:
                    embo_file_out.write(line)
                    
            embo_file_in.close()
            embo_file_out.close()
            
            os.remove(name_embo)
    
# =============================================================================
#     def three_year_Expiry():
#         current_date = datetime.now()
#         three_years_from_now = current_date + timedelta(days=3*365)
#         month = three_years_from_now.strftime('%m')
#         year = three_years_from_now.strftime('%Y')
#         return f"{month}/{year}"
# =============================================================================
    
    def contains_special_characters(input_string):
        for char in input_string:
            asscii_value = ord(char)
            if 65 <= int(asscii_value) <= 90 or 97 <= int(asscii_value) <= 122 or int(asscii_value) == 32:
                a = True
            else:
                a = False
                break
        return a
    
    def program_closed():#=============================================del extra files
        closeInput = input("Press ENTER to exit")
        sys.exit()
        print ("Closing...")  
        
    def pp_count_check(emb_count):
        count_pp = 0
        with open('config/IXIGO_PP_NUMBER.csv', 'r') as IXIGOin:
            contents = IXIGOin.readlines()
            count_pp = len(contents)
            # print(count_pp)
        IXIGOin.close()
        if count_pp<emb_count:
            return count_pp
        else:
            return 1
        
        
    file_name = glob.glob('PP*.txt')
    if len(file_name)>=1:
        for file in file_name:
            # print(file)
            with open(file, 'r') as filehandle:
                filecontents = filehandle.readlines()
                
                pp_count = pp_count_check(len(filecontents))
                
                if pp_count!=1:
                    print("PP remaining count is lest than input file , Remaining Count : "+str(pp_count))
                    program_closed()
                    
                for line1 in filecontents[1:]:
                    line = line1.split('~')
                    PP_CardNumber   = line[0].rstrip()
                    CR_CardNumber   = line[1].rstrip()
                    CR_CardNumber_first_six = CR_CardNumber[0:6]
                    
                    # print(CR_CardNumber_first_six)
                    # print(Expiry_Date_1)
                    Customer_Name   = line[2].rstrip()
                    Contact_Number  = line[3].rstrip() 
                    Mailing_Address = line[4].rstrip()
                    Title           = line[5].rstrip()
                    CardHolder_Name = line[6].rstrip()
                    logo            = line[7].strip()
                    Bin             = line[8].strip()
                    product         = line[9].strip()
                    City_PIN_1      = line[10]
                    Expiry_Date     = line[11].strip()
                    State           = line[12].rstrip() 
                    Account_Number  = line[13].rstrip()
                    Card_Action     = line[14].rstrip()
                    International   = line[15].strip()
                    Domestic        = line[16].strip()
                    CaseID          = line[17].strip()
                    Address_1       = ''
                    Address_2       = ''
                    Address_3       = ''
                    product         = product.replace('+', '_PLUS').replace(' ', '_')
                    
                    name_check = ''
                    name_check = contains_special_characters(CardHolder_Name)
                    
                    if name_check == False :
                        with open("Special_Character_Records.csv",'a') as special_chr_rec:
                            print("Special Character Found")
                            special_chr_rec.write(CardHolder_Name +','+CR_CardNumber+'\n') 
                        
                    if CR_CardNumber_first_six != '406977' and CR_CardNumber_first_six != '457036' and CR_CardNumber_first_six != '465523':
                        with open("Wrong_BIN_Records.csv",'a') as special_chr_rec:
                            print('NEW BEEN FOUND')
                            special_chr_rec.write(CardHolder_Name +','+CR_CardNumber+'\n') 
                    else:   
                        try:
                            
                            City            = City_PIN_1[:-7].rstrip()
                            PIN             = City_PIN_1[-6:].rstrip()
                        except:
                            print(f"Error in pincode in account number : {Account_Number}\n")
                            closeInput = input("Press ENTER to exit")
                            print ("Closing...")
                            
                        if logo == '301' :
                            complimentory_lounge_access = '1 per calendar quarter'
                        elif logo == '503' or logo == '306':
                            complimentory_lounge_access = '1 per calendar year'
                        elif logo == '401':
                            complimentory_lounge_access = '	2 per calendar quarter'
                        elif logo == '402' :
                            complimentory_lounge_access = '	4 per calendar quarter'
                        
                        sep             = '|'
                        
                        pp_number_1       =  pp_num(Bin)
                        
                        pp_number = pp_dup_check(pp_number_1)
                        # print(pp_number)
                        courier         =  courier_mode(PIN)[0]
                        routing_code    =  courier_mode(PIN)[1]
                        awb             =  awb_assign(courier)
                        test            =  courier[0:2].upper()
                        test            =  test.replace('SP','IP')
                        courier = courier[0:2].upper()
        
                        courier = courier.replace('DE', 'DL')
                        courier = courier.replace('SP', 'IP')
                        courier = courier.replace('BL', 'BD')
                        
                        ixigo_count = len(pp_ixigo_count)
                        pp_ixigo_remaning_count = (ixigo_Total_count - ixigo_count)
                        
                        zenith_count = len(pp_zenith_count)
                        pp_zenith_remaining_count = (ZENITH_Total_count - zenith_count )
                        
                        vetta_count = len(pp_VETTA_count)
                        pp_vetta_remaining_count = (VETTA_Total_count - vetta_count )
                        
                        refrence_num    =  courier+'AUCPP'+ awb[:11]
                        if Title == 'Mrs.':
                           Title = Title.replace('Mrs.', 'Ms.')
                           # print('Yes')
                        
                        track1_acc = '%PP/'+Title+'/'+CardHolder_Name+'//''?'
                        track2_acc = ';'+pp_number+'='+ Expiry_Date.replace('/', '20') +'?'
                        
                        # if Bin   =='465523':
                        #     product = 'VETTA'
                        #     logo    = '301'
                        
                        # elif Bin == '457036':
                        #     product = 'ZENITH'
                        #     logo    = '401'
                        
                        # elif Bin == '406977':
                        #     product = 'IXIGO'
                        #     logo    = '503'
                        
                        # print(Title.upper()+''+Customer_Name.rstrip().upper())
                        
                        with open('xxAUC_PP_NEW_ISSUE_'+str(Bin)+'_L'+str(logo)+'_'+str(product)+'_'+str(ptime)+'_EM.txtt', 'a') as pp_file:
                            pp_file.write('#0#'+CR_CardNumber+'#1#'+pp_number+'#2#'+track1_acc+'#3#'+track2_acc+'#4#'+Expiry_Date+'#5#'+Title.upper()+' '+CardHolder_Name.upper()+'#6#'+courier +'\n')
                            #print (file_pp)
                        pp_file.close()
                        
                        with open('AUC_PP_NEW_ISSUE_Data_Allocation_Report_raw.txt','a') as alloctaion_file:
                            alloctaion_file.write(Account_Number+'|'+CR_CardNumber+'|'+Title.upper()+' '+CardHolder_Name.rstrip().upper()+'|'+'|'+'|'+'|'+'|'+pp_number+'|'+'|'+'|'+'|'+refrence_num+'|'+awb+'|'+courier+'|'+routing_code+'|'+Card_Action+'|'+'|'+'|'+'1'+'|'+Bin+'|'+logo+'|'+'|'+product+'|'+Mailing_Address+'|'+Address_2+'|'+Address_3+'|'+'|'+City+'|'+State+'|'+PIN+'|'+Contact_Number+'|'+'|'+'1'+'|'+product+'|'+complimentory_lounge_access+'|'+Expiry_Date.strip()+'\n')
                           # |Account Number|Primary Card number|Customer Name|Add-on Card 1|Add-on Card 2|Add-on Card 3|Add-on Card 4|PP  Card No_Primary|PP  Card No_Addon1|PP  Card No_Addon2|PP  Card No_Addon3|Ref no.|AWB. No.|Courier|Courier Code|Card Action|Primary Count|Add-on Count|Total Cards|Bin|Logo|Gender Code|Varient|Address Line 1|Address Line 2|Address Line 3|Address Line 4|City|State Code|Postal Code|Mobile Number|Cust Unique ID|Credit Limit|Statement Date|Extention|AccountHolder_Name|Product Varient|Complimentary Lounge Access|Expiry Date 
                        with open(f"AUC-PP-NEW_ISSUE-FF_{str(product)}_.txt",'a') as ff_file:
                            ff_file.write(Account_Number+'|'+CR_CardNumber+'|'+Title.upper()+''+CardHolder_Name.rstrip().upper()+'|'+'|'+'|'+'|'+'|'+pp_number+'|'+'|'+'|'+'|'+refrence_num+'|'+awb+'|'+courier+'|'+routing_code+'|'+Card_Action+'|'+'|'+'|'+'1'+'|'+Bin+'|'+logo+'|'+'|'+product+'|'+Mailing_Address+'|'+Address_2+'|'+Address_3+'|'+'|'+City+'|'+State+'|'+PIN+'|'+Contact_Number+'|'+'|'+'|'+'|'+'|'+Customer_Name+'|'+product+'|'+complimentory_lounge_access+'|'+Expiry_Date.strip()+'\n')
                        
                        with open('AUC-PP-NEW_ISSUE-MIS_.txt','a') as mis_file:
                            mis_file.write(Account_Number+'|'+CR_CardNumber+'|'+Title.upper()+''+CardHolder_Name.rstrip().upper()+'|'+'|'+'|'+'|'+'|'+pp_number+'|'+'|'+'|'+'|'+refrence_num+'|'+awb+'|'+courier+'|'+routing_code+'|'+Card_Action+'|'+'|'+'|'+'1'+'|'+Bin+'|'+logo+'|'+'|'+product+'|'+Mailing_Address+'|'+Address_2+'|'+Address_3+'|'+'|'+City+'|'+State+'|'+PIN+'|'+Contact_Number+'|'+'|'+'|'+'|'+'|'+Customer_Name+'|'+product+'|'+complimentory_lounge_access+'|'+Expiry_Date.strip()+'\n')
                        
                        
                        with open('config/Priority_Pass_Data_AUF_BANK.csv', 'a') as pp_file_data:
                            pp_file_data.write(pp_number+','+str(ptime[0:10])  +','+Expiry_Date.replace('/', '20')+','+Account_Number+','+CardHolder_Name.rstrip()+'\n')
           
                        pp_file_data.close()
                        
                        if courier == 'BD':
                            with open('xxxAUF_Credit_Bluedart_Courier_Connection_Report_.txt', 'a') as pp_file_auf_ha:
                                 pp_file_auf_ha.write('|'+refrence_num+'|'+Title+' '+CardHolder_Name.rstrip()+'|'+Mailing_Address+'|||'+City+'|'+State+'|INDIA|'+PIN+'|'+Contact_Number+'|'+awb+'\n')
                            pp_file_auf_ha.close()
                        
                        elif courier == 'DL':
                            with open('xxxAUF_Credit_Delhivery_Courier_Connection_Report_.txt', 'a') as pp_file_auf_ha:
                                pp_file_auf_ha.write('|'+awb+'|'+refrence_num+'|'+Title+' '+CardHolder_Name.rstrip()+'|'+City+'|'+State+'|INDIA|'+Mailing_Address+'|'+PIN+'|'+Contact_Number+'|'+'50|Prepaid|500|Secure Deliverables|AU Small Finance Bank Limited M Floor, CP3-232, Industrial Area, Apparel Park, Mahal Road, Jagatpura, Jaipur, 302022|302022|AU Small Finance Bank Ltd|AU Small Finance Bank Limited AU Centre, 3rd, 5th, 6th & 7th Floor, Sunny Trade Centre, New Atish Market, Jaipur, Rajasthan  302019|True|True\n')
                            pp_file_auf_ha.close()
                        
                        else:
                            with open('xxxAUF_Credit_IndiaPost_Courier_Connection_Report_.txt', 'a') as pp_file_auf_ha:
                                pp_file_auf_ha.write('|'+awb+'|'+refrence_num+'|'+City+'|'+PIN+'|'+Mailing_Address+'|'+'|'+'|'+'|' +'|'+Contact_Number+'||'+str(awb)+'\n')
                            pp_file_auf_ha.close()
                    
                        with open('Daily_Priority_Pass_issue_Data_'+str(ptime)+'.csv', 'a') as pp_file_data:
                            if header == 0:
                                pp_file_data.write('Account_Number,Logo,Card_Number,Card_Holder_name,Card_action,Priority_Pass_Number,PP_issuance_date\n')
                            pp_file_data.write(Account_Number+','+logo  +','+CR_CardNumber+','+Title+' '+CardHolder_Name.rstrip()+','+Card_Action+','+pp_number+','+str(ptime[0:10])+'\n')
                            header+=1
                        pp_file_data.close()
                        
                        
                        
                        
                        
            filehandle.close()
    else:
        print('No input peresent')
        program_closed()
    IndiaPost_Courier = 'xxxAUF_Credit_IndiaPost_Courier_Connection_Report_.txt'
    if Path(IndiaPost_Courier).is_file():
        records_list = []
        s_no = 0
        with open('xxxAUF_Credit_IndiaPost_Courier_Connection_Report_.txt', "r") as ff, open('AUF_Credit_IndiaPost_Courier_Connection_Report_.txt', "w") as foutfile:
            foutfile.write('SNO|Barcode_Value|REFERANCE NUMBER|CITY|PINCODE|NAME|ADDRESS 1|ADDRESS 2|ADDRESS 3|ADDRESSEE EMAIL|ADDRESSEE MOBILE|SENDER MOBILE|POD REQUIRED'+'\n')
            contents = ff.readlines()
            for fi in contents:
                records_list.append(fi)
                s_no = (records_list.index(fi)+1)
                s_no = "%05d" % s_no
                if not fi.strip():
                    continue
                if fi:
                    foutfile.write(str(s_no)+fi)
        ff.close()
        foutfile.close()
        df = pd.read_csv('AUF_Credit_IndiaPost_Courier_Connection_Report_.txt',encoding= 'unicode_escape', sep='|', dtype=object)
        df.to_excel('AUF_Credit_PP_NEW_ISSUE_IndiaPost_Courier_Connection_Report_'+str(ptime)+'--'+str(s_no)+'.xlsx', 'Sheet1', index=False)
        os.remove('xxxAUF_Credit_IndiaPost_Courier_Connection_Report_.txt')
        os.remove('AUF_Credit_IndiaPost_Courier_Connection_Report_.txt')
    
    Delhivery_Courier = 'xxxAUF_Credit_Delhivery_Courier_Connection_Report_.txt'
    if Path(Delhivery_Courier).is_file():
        records_list=[]
        s_no=0
        with open('xxxAUF_Credit_Delhivery_Courier_Connection_Report_.txt',"r") as ff, open('AUF_Credit_Delhivery_Courier_Connection_Report_.txt',"w") as foutfile:
            foutfile.write('Waybill|Order No/Reference No|Consignee Name|City|State|Country|Address|Pincode|Phone/Mobile|Weight|Payment Mode|Package Amount|Product to be Shipped|Return Address|Return Pin|Seller Name|Seller Address|person_specific|address_specific'+'\n')
            contents = ff.readlines()
            for fi in contents:
                records_list.append(fi)
                s_no= (records_list.index(fi)+1)
                s_no = "%05d" % s_no
                if not fi.strip():
                    continue
                if fi:
                    foutfile.write(fi)
        ff.close()
        foutfile.close()
        
    
        df = pd.read_csv('AUF_Credit_Delhivery_Courier_Connection_Report_.txt',encoding= 'unicode_escape', sep='|', dtype=object)
        df.to_excel('AUF_Credit_PP_NEW_ISSUE_Delhivery_Courier_Connection_Report_'+str(ptime)+'--'+str(s_no)+'.xlsx', 'Sheet1',index=False)
        os.remove('xxxAUF_Credit_Delhivery_Courier_Connection_Report_.txt')
        os.remove('AUF_Credit_Delhivery_Courier_Connection_Report_.txt')
    
    Bluedart_Courier = 'xxxAUF_Credit_Bluedart_Courier_Connection_Report_.txt'
    if Path(Bluedart_Courier).is_file():
        records_list = []
        s_no = 0
        with open('xxxAUF_Credit_Bluedart_Courier_Connection_Report_.txt', "r") as ff, open('AUF_Credit_Bluedart_Courier_Connection_Report_.txt', "w") as foutfile:
            foutfile.write('Refrence Number|Customer Name|ADDRESS1|ADDRESS2|ADDRESS3|CITY|STATE|COUNTRY|PINCODE|CONTACT1|AWB'+'\n')
            contents = ff.readlines()
            for fi in contents:
                records_list.append(fi)
                s_no = (records_list.index(fi)+1)
                s_no = "%05d" % s_no
                if not fi.strip():
                    continue
                if fi:
                    foutfile.write(fi)
        ff.close()
        foutfile.close()
        df = pd.read_csv('AUF_Credit_Bluedart_Courier_Connection_Report_.txt',encoding= 'unicode_escape', sep='|', dtype=object)
        df.to_excel('AUF_Credit_PP_NEW_ISSUE_Bluedart_Courier_Connection_Report_'+str(ptime)+'--'+str(s_no)+'.xlsx', 'Sheet1', index=False)
        os.remove('xxxAUF_Credit_Bluedart_Courier_Connection_Report_.txt')
        os.remove('AUF_Credit_Bluedart_Courier_Connection_Report_.txt')    
        
        
    with open('AUC_PP_NEW_ISSUE_Data_Allocation_Report_raw.txt','r') as alloctaion_file ,open('AUC_PP_NEW_ISSUE_Data_Allocation_Report_'+ptime+'.txt','w') as alloctaion_file_out:
        alloctaion_file_out.write('Sr. No|Account Number|Primary Card number|Customer Name|Add-on Card 1|Add-on Card 2|Add-on Card 3|Add-on Card 4|PP  Card No_Primary|PP  Card No_Addon1|PP  Card No_Addon2|PP  Card No_Addon3|Ref no.|AWB. No.|Courier|Courier Code|Card Action|Primary Count|Add-on Count|Total Cards|Bin|Logo|Gender Code|Varient|Address Line 1|Address Line 2|Address Line 3|Address Line 4|City|State Code|Postal Code|Mobile Number|Cust Unique ID|Total PP Count|Product Varient|Complimentary Lounge Access|Expiry Date\n')
        contents  = alloctaion_file.readlines()
        sr_no = 1
        
        for line in contents:
            s_no = "%04d" % sr_no
            alloctaion_file_out.write(str(s_no)+'|'+line)
            sr_no+=1
            
    alloctaion_file.close()
    alloctaion_file_out.close()
    sr_no-=1
    os.remove('AUC_PP_NEW_ISSUE_Data_Allocation_Report_raw.txt')
    df = pd.read_csv('AUC_PP_NEW_ISSUE_Data_Allocation_Report_'+ptime+'.txt',encoding= 'unicode_escape', sep='|', dtype=object)
    df.to_excel('AUC_PP_NEW_ISSUE_Data_Allocation_Report-'+str(sr_no)+'_'+str(ptime)+'.xlsx', 'Sheet1', index=False)
    os.remove('AUC_PP_NEW_ISSUE_Data_Allocation_Report_'+ptime+'.txt')     
    
    ff_files = glob.glob("AUC-PP-NEW_ISSUE-FF_*")
    for f in ff_files:
        excel_convertor(f)
        
    ff_files = glob.glob("AUC-PP-NEW_ISSUE-MIS*")
    for f in ff_files:
        excel_convertor(f)    
        
    
    embo_files = glob.glob("xxAUC_PP_NEW_ISSUE_*")
    for f in embo_files:
        embo_sorting(f)
       
        
    current_directory = os.getcwd()
    excel_files = glob.glob(os.path.join(current_directory, '*MIS*.xlsx'))

    if excel_files:
        for excel_file in excel_files:
            convert_to_read_only(excel_file)
          
    
    
    
    #===============================================Batch card=================================================
    
    file_count = '.'
    
    os.chdir(file_count)
    names={}
    for fn in glob.glob('*AUC*.txtt'):
        with open(fn) as f:
            names[fn]=sum(1 for file_count in f if file_count.strip() and not file_count.startswith('~'))       
    with open('AUF_FILE_COUNT_'+str(ptime)+'.csv', 'w') as f:
        f.write('File Name,File Count,Non-Replace,Replace'+'\n')
        [f.write('{0},{1}\n'.format(key, value)) for key, value in names.items()] 
    
    from prettytable import PrettyTable
    ptx = PrettyTable()
    ptx.field_names = ["Bin", "Artwork No.","JB No.", "EMBOSS Filename","Qty", "Job setup", "Method -Ribbon/Foil","Chip Name"]
    
    for key, value in names.items():
        art_work = 'Not_Found'
        job_setup = 'Not_Found'
        dg_color = 'Not_Found'
        with open('config/batchcard.txt','r') as data_file:
            contents = data_file.readlines() 
        
            
            for line in contents[1:]:
                line_data = line.split(',')
                file_start_index = line_data[0].strip()
                file_end_index = line_data[1].strip()
                file_embo_name = line_data[2].strip()
                file_art_work = line_data[3].strip()
                file_job_setup = line_data[4].strip()
                file_dg_color = line_data[5].strip()
                
                # print(file_start_index+':'+file_end_index)
                if key[int(file_start_index):int(file_end_index)] == file_embo_name:
                    art_work = file_art_work
                    job_setup = file_job_setup
                    dg_color = file_dg_color
                    break
        
        ptx.add_row([key[17:23], art_work,"", key,value, job_setup, dg_color,""])
        
    
    ptx.align = "c"
    ptd=ptx.get_string()
    
    ptx1 = PrettyTable()
    ptx1.field_names = ["Emboss Filename","Supervisor Name","Signature", "  Date  ","  Time  ","   Remark(if any)   "]
    for key, value in names.items():
        
        ptx1.add_row([key,"","","","",""])
    ptx1.align = "c"
    ptd1=ptx1.get_string()
    qty1 = sum(names.values())
    datedmy = str(x)[8:10]+'-'+str(x)[5:7]+'-'+str(x)[0:4]
    with open('AUF_CREDIT_PP_NEW_ISSUE_BATCHCARD_'+str(ptime)+'.dat', 'w') as file:
      file.write('                                                '+'BANKING PERSONALISATION BATCH CARD'+'                              '+'Date: '+ptime[0:10]+'\n')
      file.write('                                                       AUC-'+ptime[0:10]+'-'+batch_number+'                                    '+'AUF CREDIT PROJECT(DI)\n\n')
      file.write(str(ptd))
      file.write('\n')
      file.write('TOTAL BATCH QUANTITY:'+str(sum(names.values()))+'\n')
      file.write('                                                   Data upload on Machine\n')
      file.write(str(ptd1))
      file.write('\n\n')
      file.write('PRP: 20.1                                              Rev No: 3.1                                            Date: 20-May-21\n')
      file.write('SEC-3: INTERNAL                                        Owner: Quality Control                                 Status: Issued\n\n')
      file.write('                                                       Page: 1 of 1')
         
    
    
    from fpdf import FPDF
    class PDF(FPDF):
        def header(self):
            self.image('config/colorplast_logo.png', x = 10, y = 5, w = 50, h = 10, type = '', link = '')
            self.ln(8)
    pdf = PDF()
    pdf.add_page('L')
    pdf.set_font("Courier", size=8)
    f = open('AUF_CREDIT_PP_NEW_ISSUE_BATCHCARD_'+str(ptime)+'.dat', "r")
    
    for x in f:
        pdf.cell(50, 5, txt = x, ln=True, align = 'L')
        #pdf.image('config/colorplast_logo.png', x = 10, y = 5, w = 50, h = 20, type = '', link = '')
        #pdf.cell(ln=10)
    
    pdf.output('AUF_CREDIT_PP_NEW_ISSUE_BATCHCARD_'+ptime[0:10]+'_'+batch_number+'.pdf')
    f.close()

    
    
    
    #---------------------------------DELETION LOG-----15.2.2----------------------
    from prettytable import PrettyTable
    ptx = PrettyTable()
    ptx.field_names = (["Client Name", "File Name","File Deletion Date", "Dispatch Date","Data Admin. Name", "Data Admin. Sign.","IT Person Name", "IT Person Sign."])




    for key,value in names.items():
     
        ptx.add_row(["AUF Bank", key, "", "", "Data Team", "", "IT Team", ""'\n'])
    ptx.align = "c"
    ptd=ptx.get_string()

    with open('AUF_CREDIT_PP_NEW_ISSUE_DELETION_LOG_'+datedmy+'.dat', 'w') as file:
      file.write('\n\n                                                                        SEC-IS   |   02.02\n')
      file.write('                                                                        Rev.No.  |   3.0                                                   '+'AUF BANK\n')
      file.write('                                                                        Date :   |   15-Jul-18\n')
      file.write('                                                                 '+'( ***** DATA DELETION LOG ***** )''\n\n\n\n')
      file.write('|  Data Receiving Date - '+datedmy+'  |  '+'  Data Servier - DPP Server/MX Machine  |   '+'  Batch Qty. - '+str(sum(names.values()))+'   |  '+'  Batch No. - AUF-'+str(x)[8:10]+'-'+str(x)[5:7]+'-'+str(x)[0:4]+'--'+batch_number+'  |  '+'Status -             '+'|''\n\n\n')

      file.write(str(ptd))
      file.write('\n\n\n\n')

    class PDF(FPDF):
        def header(self):
            self.image('config/colorplast_logo.png', x = 10, y = 5, w = 50, h = 10, type = '', link = '')
            self.ln(8)

    pdf = PDF()
    pdf.add_page('L')
    pdf.set_font("Courier", size=7)
    f = open('AUF_CREDIT_PP_NEW_ISSUE_DELETION_LOG_'+datedmy+'.dat', "r")

    for x in f:
        pdf.cell(50, 5, txt = x, ln=True, align = 'L')
    pdf.output('AUF_CREDIT_PP_NEW_ISSUE_DELETION_LOG_'+datedmy+'.pdf')
    f.close()

    #---------------------------------DELETION LOG ENDS ---------------------------
    
    
    path = ('.')
    ext = "txtt"
    
    for f in os.listdir(path):
        fpath = os.path.join(path, f)
    
        if os.path.isfile(fpath) and fpath.endswith(ext):
            time = datetime.fromtimestamp(os.path.getctime(fpath)).strftime("%d-%m-%Y-%H%M%S----"+str(sum(names.values())))
            name='AUF_CREDIT_PP_NEW_ISSUE_'+time
            os.makedirs(os.path.join(path, name), exist_ok=True)
            os.replace(fpath, os.path.join(path, name, f))
   
    from distutils.dir_util import copy_tree
    import os


    os.chdir(path) 
    fromDirectory = os.getcwd()
    source_directory_name = name
    source_directory = os.path.join(fromDirectory, source_directory_name)
    target_directory = "A:/Sdrive/LIVE/AU_Credit/"+name
    copy_tree(source_directory, target_directory)
    
    del_filetxt = glob.glob("*.txt")
    for de in del_filetxt:   
        os.remove(de)
    
    with open('config/ZENITH_PP_NUMBER_USING.csv', 'r') as zenithin, open('config/ZENITH_PP_NUMBER.csv', 'w') as zenithout:
        contents = zenithin.readlines()
        zenithout.writelines(contents)
    zenithin.close()
    zenithout.close()
    
    with open('config/VETTA_PP_NUMBER_USING.csv', 'r') as vettain, open('config/VETTA_PP_NUMBER.csv', 'w') as vettaout:
        contents = vettain.readlines()
        vettaout.writelines(contents)
    vettain.close()
    vettaout.close()
    
    with open('config/IXIGO_PP_NUMBER_USING.csv', 'r') as IXIGOin, open('config/IXIGO_PP_NUMBER.csv', 'w') as IXIGOout:
        contents = IXIGOin.readlines()
        IXIGOout.writelines(contents)
    IXIGOin.close()
    IXIGOout.close()
    
   
    try:
        os.remove('config/ZENITH_PP_NUMBER_USING.csv')
        os.remove('config/VETTA_PP_NUMBER_USING.csv')
        os.remove('config/IXIGO_PP_NUMBER_USING.csv')
    except Exception as e:
        print(e)
     
    with open('config/BD_AWB.txt', 'r') as bdfin:
        contents= bdfin.readlines()
        bdremainingnew = len(contents)
    bdfin.close()
    with open('config/DL_AWB.txt', 'r') as dtfin:
        contents= dtfin.readlines()
        dtremainingnew = len(contents)    
    dtfin.close()
    with open('config/IP_AWB.txt', 'r') as ipfin:
        contents= ipfin.readlines()   
        ipremainingnew = len(contents)
    ipfin.close()
    
    
    
    with open('AWB_REMAINING_COUNT_NEW.csv','w') as awb_count_new:
        awb_count_new.write('BD   : ' + str(bdremainingnew)+'\n'+'DL : '+str(dtremainingnew)+'\nIP   : '+str(ipremainingnew) )
        print('BD       |     DL         |      IP   ')
        print(str(bdremainingnew)+'    |     '+str(dtremainingnew)+'     |       '+str(ipremainingnew) )
    awb_count_new.close()
    
    with open('PP_REMAINING_COUNT.csv','w') as pp_count_new:
        pp_count_new.write('PP_IXIGO  : ' + str(pp_ixigo_remaning_count)+'\n'+'PP_ZENITH : '+str(pp_zenith_remaining_count)+'\nPP_VETTA   : '+str(pp_vetta_remaining_count))
        print('PP_IXIGO             |',str(pp_ixigo_remaning_count))
        print('PP_ZENITH            |',str(pp_zenith_remaining_count))
        print('PP_VETTA             |',str(pp_vetta_remaining_count))
        
        # print(str(pp_remaning_count))
        
    pp_count_new.close()

    
    current_date = datetime.now()
    dt1=current_date.strftime("%Y-%m-%d-%H%M%S---")


    for f in os.listdir(path):
        fpath = os.path.join(path, f)
        
        if os.path.isfile(fpath) and (fpath.endswith('xlsx') or fpath.endswith('pdf') or fpath.endswith('dat') or fpath.endswith('csv')):
            # time = datetime.fromtimestamp(os.path.getctime(fpath)).strftime("%d-%m-%Y-%H%M%S---"+str(qty))
            time = datetime.fromtimestamp(os.path.getctime(fpath)).strftime(dt1+str(qty1).rjust(4,'0'))
            name='AUF_CREDIT_PP_NEW_ISSUE_FF_MIS_'+time
            os.makedirs(os.path.join(path, name), exist_ok=True)
            os.replace(fpath, os.path.join(path, name, f))


    from distutils.dir_util import copy_tree
    
    with open('config/Output_file_location.csv','r') as file:
        content = file.readlines()
        folder_location = content[0].split(',')[1].strip()
        
    os.chdir(path) 
    fromDirectory = os.getcwd()
    aaa = fromDirectory+'/'+name
    toDirectory = folder_location 
    bbb = toDirectory+name
    copy_tree(aaa, bbb)
    
    
    
    

except Exception as e :
    print("Error:",e)
print('Version - 1.0.5')   
closeInput = input("Press ENTER to exit")

print ("Closing...")    
