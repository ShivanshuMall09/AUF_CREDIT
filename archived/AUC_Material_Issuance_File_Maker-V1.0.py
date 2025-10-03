import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import re
from datetime import datetime
import glob

print('\n!! Welcome to Colorplast Systems Co. !!\n')
ptime = datetime.now().strftime("%A, %d.%m.%Y - %H-%M-%S")
print("Today's : " + ptime)
print()

def process_excel_file(file_to_process):
    try:
        input_dir_path = os.getcwd()
        
        # Construct the full file path for the selected file
        input_file_path = os.path.join(input_dir_path, file_to_process)
        # Extract date and count from the file name (e.g., "03.01.2025_352")
        match = re.search(r"(\d{2}\.\d{2}\.\d{4})_(\d+)", file_to_process)
        if match:
            date_part = match.group(1)  # Extract date (e.g., "03.01.2025")
            count_part = match.group(2)  # Extract count (e.g., "352")
            output_file_name = f"AUC_Material_Issuance_Detail_{date_part}--{count_part}.xlsx"
        else:
            print("Date and count could not be extracted from the file name.")
            return

        # Read the input Excel file
        input_df = pd.read_excel(input_file_path)

        # Group by "Card Body A/W Ref.No." (or "ARTWORK NO.") and sum the "Quantity"
        grouped_df = input_df.groupby(["ARTWORK NO.", "BIN"], as_index=False).agg({
            "Quantity": "sum"  # Sum the "Quantity"
        })

        # Define output headers
        output_headers = [
            "Item Requested",
            "Card Body A/W Ref.No.",
            "BIN",
            "Req Quantity",
            "Issued Quantity",
            "Request By Name",
            "Request By Sign",
            "Date & Time",
            "Issued / Name",
            "Issued / Sign",
            "Date & Time1",
            "Remark If any."
        ]

        # Create the output DataFrame with the required headers
        output_df = pd.DataFrame(columns=output_headers)

        # Populate the output DataFrame
        output_df["Item Requested"] = ["Card"] * len(grouped_df)  # Fixed value "Card"
        output_df["Card Body A/W Ref.No."] = grouped_df["ARTWORK NO."]  # Map from grouped data
        output_df["BIN"] = grouped_df["BIN"]  # Map from grouped data
        output_df["Req Quantity"] = grouped_df["Quantity"]  # Map summed "Quantity"

        # Add empty columns for other headers
        output_df["Issued Quantity"] = ''
        output_df["Request By Name"] = ''
        output_df["Request By Sign"] = ''
        output_df["Date & Time"] = ''
        output_df["Issued / Name"] = ''
        output_df["Issued / Sign"] = ''
        output_df["Date & Time1"] = ''
        output_df["Remark If any."] = ''

        # Compute the sum of the "Quantity" column
        total_quantity = grouped_df["Quantity"].sum()

        # Save the output file in the same directory as the input file
        output_file_path = os.path.join(input_dir_path, output_file_name)
        output_df.to_excel(output_file_path, index=False)

        # Open the generated Excel file and apply formatting
        wb = load_workbook(output_file_path)
        ws = wb.active

        # Set sheet view to 85%
        ws.sheet_view.zoomScale = 85

        # Make the first row gray
        gray_fill = PatternFill(start_color="BEBEBE", end_color="BEBEBE", fill_type="solid")
        for cell in ws[1]:
            cell.fill = gray_fill

        # Freeze the first row
        ws.freeze_panes = "A2"

        # Add borders to all data cells
        border = Border(left=Side(border_style="thin"),
                        right=Side(border_style="thin"),
                        top=Side(border_style="thin"),
                        bottom=Side(border_style="thin"))

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = border
                # Center align all cells
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # Auto-fit column width based on the maximum length of data in each column
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter  # Get column name (e.g., 'A', 'B', etc.)
            for cell in col:
                try:
                    # For both string and numeric values, get the max length
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            # Adjust width with extra padding for readability
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        # Save the formatted output
        wb.save(output_file_path)

        print(f"\nOutput file saved at: {output_file_path}")
        print(f"\nTotal Data Count : {total_quantity}")

    except Exception as e:
        print(f"An error occurred: {e}")


input_file = glob.glob('AUF_CREDIT_BATCHCARD*.xlsx')
process_excel_file(input_file[0])

print("\n!! Required Files Generated !!")
closeInput = input("Press ENTER to exit")
print("Closing...")
