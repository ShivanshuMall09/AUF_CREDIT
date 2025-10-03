"""
Created on Thu Nov  5 16:22:49 2020

@author: rajeev.jadaun
AUF_CREDIT_ProcessV.17.1.23.16.10.py:07-05-2025[Adding last 4 of pan in scanning file]
AUF_CREDIT_ProcessV.17.1.23.16.9.py:07-05-2025[For Kosmo we have to assign different awb on the normal and addon]
AUF_CREDIT_ProcessV.17.1.23.16.8.py:30-05-2025[1) Making FPA File For tool scanning tool and auto transfer to the particular location
                                               2) Making Material issuance file for the MIS team
AUF_CREDIT_ProcessV.17.1.23.16.7.py:25-04-2025[1- Making Scanning Card file for the ff tapping tool
                                                2- changing the value of card type code which is given on team NW for new issue and so on.]
AUF_CREDIT_ProcessV.17.1.23.16.6.py:17-04-2025[Adding the gender code value in the allocation file]
AUF_CREDIT_ProcessV.17.1.23.16.5.py:11-04-2025[Fixing the problem in Replace]
AUF_CREDIT_ProcessV.17.1.23.16.4.py:10-04-2025[G1 G2 merger for the all the logo except 101,102,401]
AUF_CREDIT_ProcessV.17.1.23.16.3.py:07-04-2025[Addition of logo 421 (AU_Traverse)  bin 457036 for the generation of PP_card
                                               Addition of product cod eon the basis of the product]
AUF_CREDIT_ProcessV.17.1.23.16.2.py:27-02-2025[version name in bacthcards]
AUF_CREDIT_ProcessV.17.1.23.16.1.py:26-02-2025[removing chip name column from the batchcard]
AUF_CREDIT_ProcessV.17.1.23.16.py:14-01-2025[.. replacing it occur due to padding fault]
AUF_CREDIT_ProcessV.17.1.23.14.py:05-12-2024[Adding bin 653023 and logo 910 and 912]
AUF_CREDIT_ProcessV.17.1.23.13.py:24-10-2024[Courier connection file Excel visible false as per customer request so after entring the [password anyone can see the data]]
AUF_CREDIT_ProcessV.17.1.23.12.py:17-10-2024[password proteected courier summary]
AUF_CREDIT_ProcessV.17.1.23.10.py:28-09-2024[Generating courier summary file]
AUF_CREDIT_ProcessV.17.1.23.9.py:27-08-2024[Stopping the generation of 401 logo pp card now onwards only 402 logo pp will be generate]
AUF_CREDIT_ProcessV.17.1.23.7.py:27-08-2024[adding two column in the allocation file for the 4th line embossing]
AUF_CREDIT_ProcessV.17.1.23.6.py:21-08-2024[CRF_NUMBER:-CRF_CP20082024][Bank will pass the plastic id “0000000015” in card data file from field position “403” with field length “10” for all retail cards where 4th line embossing default will be as “Field Club”]
AUF_CREDIT_ProcessV.17.1.23.5.py:16-08-2024[Chnages in batch card and making pp file also read only]
AUF_CREDIT_ProcessV.17.1.23.4.py:21-06-2024[Changing the Rto address in the delhivery courier file on the request of aditya and mailed ]
AUF_CREDIT_ProcessV.17.1.23.3.py:30-05-2024[Trasfering the priority pass file, processing log,special character file ,AUF_File_DupeCheck file and data allocation file automatically ]
AUF_CREDIT_ProcessV.17.1.23.1.py:21-05-2024[removing vetta pp generation (changed 465523 to 465523_1)]
AUF_CREDIT_ProcessV.17.1.22.3.py:20-05-2024[Changing the productvarient AUF_BANK_LIT_CREDIT_CARD to LIT]
AUF_CREDIT_ProcessV.17.1.22.2.py:17-05-2024[fixxing the qty on FF-MIS folder]
AUF_CREDIT_ProcessV.17.1.22.py:14-05-2024[making ff-mis merge file , adding filename in rejection]
AUF_CREDIT_ProcessV.17.1.21.py:29-04-2024[adding for total card count for Zenith_Plus or LIT]
AUF_CREDIT_ProcessV.17.1.20.py:23-04-2024[Previous the vlaue of name for special chareater check is [54-130] but on the customer request [54:82]]
AUF_CREDIT_ProcessV.17.1.19.py:11-04-2024[Fixxing the bug character embo name]
AUF_CREDIT_ProcessV.17.1.16.py:11-04-2024[Rejectting specail character embo name]
AUF_CREDIT_ProcessV.17.1.15.py:21-03-2024[Adding chip name module in batchcard]
AUF_CREDIT_ProcessV.17.1.14.py:21-02-2024[1-Making ff file Read only on the request by aamir]
AUF_CREDIT_ProcessV.17.1.13.py:08-01-2024[Awb allocation issue is resolve ]
AUF_CREDIT_ProcessV.17.1.12.py:04-01-2024[Replacing the special character from the input file]
AUF_CREDIT_ProcessV.17.1.11.py:03-01-2024[Resolving the logo problem with abcpro ]
AUF_CREDIT_ProcessV.17.0.10.py:02-01-2024[Stoping the generation of pp file for the 306 ]
AUF_CREDIT_ProcessV.17.0.9.py:30-10-2023[ AWB Issue is Resolve ]
AUF_CREDIT_ProcessV.17.0.9.py:30-10-2023[EMBO CHECKER FUNCTION IS DISABLE ON THE REQUEST of AAMIR SIR]
AUF_CREDIT_ProcessV.17.0.8.py:28-10-2023[FIXXING THE ISSUE IN REPLACE FF IS RESOLVED ]
AUF_CREDIT_ProcessV.17.0.7.py:25-10-2023[ . PROBLEM IN EMBO IS RESOLVED ]
AUF_CREDIT_ProcessV.17.0.6.py:20-10-2023[Zenith plus issue is resolve ]
AUF_CREDIT_ProcessV.16.0.5.py:14-10-2023[Adding timmer that can automatically closed the py after 12 hours]
AUF_CREDIT_ProcessV.16.0.4.py:12-04-2023[Sorting order in pp as per card type]
AUF_CREDIT_ProcessV.16.0.2.py:12-04-2023[Configurable sorting order in config]
AUF_CREDIT_ProcessV.16.0.1.py:12-04-2023[Merging all embo and ff files]  
AUF_CREDIT_ProcessV.15.5.2.py:12-04-2023[Crash Issue ]  
AUF_CREDIT_ProcessV.15.5.0.py:16-02-2023[awb allocation for new add-on only ]  
  Rules : 1. Primary + multiple add-on  : Same AWB
  Rules : 2. multiple add-on only cases : Diffrent AWB
  Rules : 3. primary with same account number cases : Diffrent AWB
  TOTAL CARD COUNTS : HOw many cards inserted in single enevelope irresspective with the no of mailers
AUF_CREDIT_ProcessV.15.4.3.py:14-02-2023[error for dupe check file within 15 days  ]  
AUF_CREDIT_ProcessV.15.3.3.py:02-01-2023 : mailer_name_acc = mailer_name1_acc
AUF_CREDIT_ProcessV.15.2.3.py:28-07-2022[Data deletion form]
AUF_CREDIT_ProcessV.15.2.2.py:23-07-2022[Delhivery Excel generation] 
AUF_CREDIT_ProcessV.15.2.1.py:19-07-2022[handeled special character in input file] 
AUF_CREDIT_ProcessV.15.1.1.py:1-07-2022[Progress report in console, error msg handling without closuing the exe]
AUF_CREDIT_ProcessV.14.1.1.py:18-06-2022[replace file handling] 
AUF_CREDIT_Process_V.13.3.1.py:01-06-2022[add a function to check the pp dp check]
AUF_CREDIT_Process_V.13.2.3.py:01-06-2022[adding artwork for G0 product for all ALTURA varients] 
AUF_CREDIT_Process_V.13.2.2.py:06-05-2022[replace varient name from CII-Young Indian to Zenith_CII_Young_Indian ] 
AUF_CREDIT_Process_V.13.2.1.py:06-05-2022[adding artwork for  new bin 483974,483976,483894, also
                                          adding aler for bin 483974] 
AUF_CREDIT_Process_V.13.1.2.py:06-05-2022[change pp expiry 3 years from date of processing] 
AUF_CREDIT_Process_V.13.1.1.py:06-05-2022[adding artwork for EM_B466505_L102_G0] 
AUF_CREDIT_Process_V.13.1.1.py:06-05-2022[adding artwork for EM_B466505_L102_G0] 
AUF_CREDIT_Process_V.13.1.0.py:21-04-2022[new bin addition]
AUF_CREDIT_Process_V.12.1.0.py:21-04-2022[Removing add on pp of vetta varient as per customer requirement]
AUF_CREDIT_ProcessV.11.2.0.py:21-12-2021[Batch card info change(artwork)]
AUF_CREDIT_ProcessV.11.3.0.py:11-01-2022[Full automation]
AUF_CREDIT_ProcessV.11.2.0.py:21-12-2021[Batch card info change(artwork)]
         


"""
from openpyxl.utils import get_column_letter
import shutil
import os
import time
import glob
from openpyxl.styles import PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import win32com.client as win32
import random
import datetime
import hashlib
from pathlib import Path
from datetime import date
import pandas as pd
from babel.numbers import format_currency
import traceback
from openpyxl import load_workbook
import importlib.machinery
import importlib.util
import warnings

# Suppress FutureWarnings
warnings.simplefilter(action='ignore', category=FutureWarning)


try:
    print("AUF_CREDIT_ProcessV.17.1.23.16.10.exe running!!!!")
    print("DATA PROCESSING STARTS.....")
    #del_file1 = glob.glob("*.txt*")
    cwd = os.getcwd()
    read_only_flag = ''
    primary_card = []
    secondary_card = []
    places = []
    cardnumber_list = []
    bin_list = []
    # cardtype_list = []
    # records_list=[]
    records_list_acc = []
    ps_list = []
    ps1_list = []
    pps_list = []
    p_acc_card_count = []
    s_acc_card_count = []
    s_acc_card_count1 = []
    pp_dup_check = []
    bd_num = 0
    dl_num = 0
    ip_num = 0
    vetta_pp = 0
    zenith_pp = 0
    pp_num =[]
    awb_number = 0
    awb = 0
    fourth_line_logo_list =[]
    pct_id_val = []
    routing_code = ''
    courier = ''
    row = []
    account_number_list = []
    pri_account_number = []
    add_on_account_number = []
    pp_list = []
    records_list = []
    s_no = 0
    
    total_primary_count = 0
    total_pp_count = 0
    count = 0
    file_count_num =0
    
    start_time = time.time()
    x = (start_time )
    ts = datetime.datetime.now()
    ptime_all = ts.strftime("%d.%m.%Y_%H%M%S")
    ptime = ts.strftime("%d.%m.%Y")
    ptime_1 = ts.strftime("%d.%m.%Y")
    datedmyd = ts.strftime("%d-%m-%Y")
    print(ptime)
    print(ptime)
    file_check = glob.glob('L2CDZU*')
    if len(file_check)==0:
        print("No Input Files present")
    else:
        
    #---------------------------FEES --------------------------------------------------------   
        with open('config/Fees_console.csv','r') as data_file_pct:
            contents = data_file_pct.readlines()
            for line in contents:
                line = line.split(',')
                # print(line[1])
                pct_id_val.append(line[1])
        data_file_pct.close()

        def process_excel_file(file_to_process):
            try:
                import re
                input_dir_path = os.getcwd()
                input_file_path = os.path.join(input_dir_path, file_to_process)
                # Extract date and count from the file name (e.g., "03.01.2025_352")
                match = re.search(r"(\d{2}\-\d{2}\-\d{4})_(\d+)", file_to_process)
                if match:
                    date_part = match.group(1)  # Extract date (e.g., "03.01.2025")
                    count_part = match.group(2)  # Extract count (e.g., "352")
                    output_file_name = f"AUC_Material_Issuance_Detail_{date_part}--{count_part}.xlsx"
                else:
                    print("Date and count could not be extracted from the file name.")
                    return

                # Read the input Excel file
                input_df = pd.read_excel(input_file_path)

                # Group by "Card Body A/W Ref.No." (or "ARTWORK NO.") and sum the "Quantity"
                grouped_df = input_df.groupby(["ARTWORK NO.", "BIN"], as_index=False).agg({
                    "QTY": "sum"  # Sum the "Quantity"
                })
                # print(grouped_df)
                # Define output headers
                output_headers = [
                    "Item Requested",                                                                                                
                    "Card Body A/W Ref.No.",
                    "BIN",
                    "Req Quantity",
                    "Issued Quantity",
                    "Request By Name",
                    "Request By Sign",
                    "Date & Time",
                    "Issued / Name",
                    "Issued / Sign",
                    "Date & Time1",
                    "Remark If any."
                ]

                # Create the output DataFrame with the required headers
                output_df = pd.DataFrame(columns=output_headers)

                # Populate the output DataFrame
                output_df["Item Requested"] = ["Card"] * len(grouped_df)  # Fixed value "Card"
                output_df["Card Body A/W Ref.No."] = grouped_df["ARTWORK NO."]  # Map from grouped data
                output_df["BIN"] = grouped_df["BIN"]  # Map from grouped data
                output_df["Req Quantity"] = grouped_df["QTY"]  # Map summed "Quantity"

                # Add empty columns for other headers
                output_df["Issued Quantity"] = ''
                output_df["Request By Name"] = ''
                output_df["Request By Sign"] = ''
                output_df["Date & Time"] = ''
                output_df["Issued / Name"] = ''
                output_df["Issued / Sign"] = ''
                output_df["Date & Time1"] = ''
                output_df["Remark If any."] = ''

                # Compute the sum of the "Quantity" column
                total_quantity = grouped_df["QTY"].sum()

                # Save the output file in the same directory as the input file
                output_file_path = os.path.join(input_dir_path, output_file_name)
                output_df.to_excel(output_file_path, index=False)

                # Open the generated Excel file and apply formatting
                wb = load_workbook(output_file_path)
                ws = wb.active

                # Set sheet view to 85%
                ws.sheet_view.zoomScale = 85

                # Make the first row gray
                gray_fill = PatternFill(start_color="BEBEBE", end_color="BEBEBE", fill_type="solid")
                for cell in ws[1]:
                    cell.fill = gray_fill

                # Freeze the first row
                ws.freeze_panes = "A2"

                # Add borders to all data cells
                border = Border(left=Side(border_style="thin"),
                                right=Side(border_style="thin"),
                                top=Side(border_style="thin"),
                                bottom=Side(border_style="thin"))

                for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                    for cell in row:
                        cell.border = border
                        # Center align all cells
                        cell.alignment = Alignment(horizontal="center", vertical="center")

                # Auto-fit column width based on the maximum length of data in each column
                for col in ws.columns:
                    max_length = 0
                    column = col[0].column_letter  # Get column name (e.g., 'A', 'B', etc.)
                    for cell in col:
                        try:
                            # For both string and numeric values, get the max length
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass
                    # Adjust width with extra padding for readability
                    adjusted_width = (max_length + 2)
                    ws.column_dimensions[column].width = adjusted_width

                # Save the formatted output
                wb.save(output_file_path)

                print(f"\nOutput file saved at: {output_file_path}")
                print(f"\nTotal Data Count : {total_quantity}")

            except Exception as e:
                print(f"An error occurred: {e}")
        
        def convert_to_read_only(file_path):
            password = 'your_password_here' 
            global read_only_flag
            
            if read_only_flag == 'NO':
                return
            else:
                try:
                    wb = load_workbook(file_path)

                    # Protect all sheets with password
                    for sheet in wb.sheetnames:
                        ws = wb[sheet]
                        ws.protection.sheet = True
                        ws.protection.password = password
                    
                    # Set sheet zoom to 85%
                    for sheet in wb.sheetnames:
                        ws = wb[sheet]
                        ws.sheet_view.zoomScale = 85
                    
                    # Auto-fit all columns
                    for sheet in wb.sheetnames:
                        ws = wb[sheet]
                        for col in range(1, ws.max_column + 1):
                            column_letter = get_column_letter(col)
                            max_length = 0
                            for cell in ws[column_letter]:
                                try:
                                    if len(str(cell.value)) > max_length:
                                        max_length = len(cell.value)
                                except:
                                    pass
                            # Set column width to accommodate the maximum content length
                            ws.column_dimensions[column_letter].width = (max_length + 2) * 1.2
                    
                    # Save the modified workbook
                    wb.save(file_path)
                except Exception as e:
                    print(f"Error converting '{file_path}': {e}")
        
        # sorting defination for EMBOSSA file #-----------------------------------------------------------------------------------
        def my_sort_emb_courier(line):
            # print(line)
            line_fields = line.strip().split('|')
            couriersort = line_fields[-2]
           
            with open('config/Sorting_Courier.txt' , 'r') as sorting_data_file:
                sorting_contents = sorting_data_file.readlines()
                for line in sorting_contents:
                    data = line.split(':')
                    if data[0]==couriersort:
                        # print(line)
                       
                        return data[1]
                        break
            return '0'  
        
        def my_sort_emb_cardtype(line):
            line_fields = line.strip().split('|')
            cardtypesort = line_fields[-1]
           
            with open('config/Sorting_CardType.txt' , 'r') as sorting_data_file:
                sorting_contents = sorting_data_file.readlines()
                for line in sorting_contents:
                    data = line.split(':')
                    if data[0]==cardtypesort:
                        # print(line)
                        return data[1]
                        break
            return '0'     
        
        def pp_sort_emb_courier(line):
            # print(line)
            line_fields = line.strip().split('#')
            couriersort = line_fields[-3]
           
            with open('config/Sorting_Courier.txt' , 'r') as sorting_data_file:
                sorting_contents = sorting_data_file.readlines()
                for line in sorting_contents:
                    data = line.split(':')
                    if data[0]==couriersort:
                        # print(line)
                       
                        return data[1]
                        break
            return '0'  
        
        
        def pp_sort_emb_cardtype(line):
            # print(line)
            line_fields = line.strip().split('#')
            couriersort = line_fields[-1]
            with open('config/Sorting_CardType.txt' , 'r') as sorting_data_file:
                sorting_contents = sorting_data_file.readlines()
                for line in sorting_contents:
                    data = line.split(':')
                    if data[0]==couriersort:
                        # print(line)
                       
                        return data[1]
                        break
            return '0' 
            # return couriersort
        # sorting defination for FF_MIS file#-----------------------------------------------------------------------------------
        def my_sort_ff_courier(line):
            line_fields = line.strip().split('|')
            couriersort = (line_fields[14])
            log_write(couriersort)
            with open('config/Sorting_Courier.txt' , 'r') as sorting_data_file:
                sorting_contents = sorting_data_file.readlines()
                for line in sorting_contents:
                    data = line.split(':')
                    if data[0]==couriersort:
                        # print(line)
                        return data[1]
                        break
            return '0' 
    
        def my_sort_ff_cardtype(line):
            line_fields = line.strip().split('|')
            cardtype = (line_fields[-2])
            log_write(cardtype)
            with open('config/Sorting_CardType.txt' , 'r') as sorting_data_file:
                sorting_contents = sorting_data_file.readlines()
                for line in sorting_contents:
                    data = line.split(':')
                    if data[0]==cardtype:
                        # print(line)
                        return data[1]
                        break
            return '0'
        
        def my_sort_pp_cardtype(line):
            line_fields = line.strip().split('|')
            couriersort = (line_fields[-1])
            log_write(couriersort)
            with open('config/Sorting_CardType.txt' , 'r') as sorting_data_file:
                sorting_contents = sorting_data_file.readlines()
                for line in sorting_contents:
                    data = line.split(':')
                    if data[0]==couriersort:
                        # print(line)
                        return data[1]
                        break
            return '0'
            # return couriersort
        
        def my_sort_pp_courier(line):
            line_fields = line.strip().split('|')
            couriersort = (line_fields[-2])
            log_write(couriersort)
            with open('config/Sorting_Courier.txt' , 'r') as sorting_data_file:
                sorting_contents = sorting_data_file.readlines()
                for line in sorting_contents:
                    data = line.split(':')
                    if data[0]==couriersort:
                        # print(line)
                        return data[1]
                        break
            return '0' 
        
# =====================log writing========================================================
        def log_write(msg):
            with open('log.dat','a') as log_data:
                log_data.write(msg+'\n')
                
            log_data.close
# =============================================================================
        
      #-----------------------PP DUP CHECK------------------------------------------------------------   
        def pp_dup_check(number):
            with open('./config/AUF_CREDIT_PP_CONSOLE.txt','r') as file_read:
                contents = file_read.readlines()
                new_contents = [x[:-1] for x in contents]
                if number in new_contents:
                    print("PP duplicate number found : "+number)
                elif number in pp_num:
                    print("PP duplicate number found : "+number)
                
                else:
                    #    ("not found")
                    pp_num.append(number)
                    
            file_read.close()
            return number
#=============================================(.) Remover =============================================
     
        def embo_name_checker(embo_dataline):
            a = embo_dataline
            b = a[54:130]
            if b.__contains__('.'):
                b= b.replace('.',' ')
                d = " ".join(b.split())
                d = d.ljust(76)
                f = a[:54]+d+a[130:]
                # print(d)
                f = a[:54]+d+a[130:]
                # print(d)
                return f
            else:
                return a
                
#=============================================Fees-Logic=============================================                
        def fees(logo,pct):   
            joining_fees = ""
            annual_fees = ""
            with open("./config/Fees_console.csv","r") as data_read:
                contents = data_read.readlines()
                #print(str(logo)+"-"+str(pct))
                
                for line in contents:
                    line = line.split(",")
                    logo_id = line[0]
                    pct_id = line[1]
                    if (logo_id == logo) and (pct_id == pct):
                        joining_fees = line[2]
                        annual_fees = line[3].rstrip()
                        break
                    
                return(joining_fees,annual_fees)
    
        def program_closed():#=============================================del extra files
            closeInput = input("Press ENTER to exit")
            print ("Closing...")  
            
 
        
        def courier_assign(pincode_1):#=============================================courier_assign
            with open('config/PINCODE_MASTER.csv', 'r') as pin:
                routing_codewe = pin.readlines()
                flag = 0
                courier_1 = ''
                routing_code_1 = ''
                # customer_code = ''
                for line in routing_codewe:
                    pincodedata = line.split(',')
                    if pincode_1 != '':
                        # and pincodedata[4]=='BLUEDART':
                        if pincode_1 in pincodedata[0]:
                            flag = 1
    
                            courier_1 = pincodedata[2].rstrip()
                            routing_code_1 = pincodedata[1].rstrip()
                            break
                if flag == 0:
                    courier_1 = 'Indiapost'
                    routing_code_1 = 'BNPL-921-413'
                    # courier_1 = str(courier_1)
            pin.close()
            return(courier_1,routing_code_1)
        
        def jdtodatestd(jdate):#=============================================courier_assigndate converson
            jdate = str(jdate)
            datestd = datetime.datetime.strptime(jdate, '%Y%j').date()
            datestd = str(datestd)
        
            if datestd[5:7] == '01':
                month = 'Jan'
            elif datestd[5:7] == '02':
                month = 'Feb'
            elif datestd[5:7] == '03':
                month = 'Mar'
            elif datestd[5:7] == '04':
                month = 'Apr'
            elif datestd[5:7] == '05':
                month = 'May'
            elif datestd[5:7] == '06':
                month = 'Jun'
            elif datestd[5:7] == '07':
                month = 'Jul'
            elif datestd[5:7] == '08':
                month = 'Aug'
            elif datestd[5:7] == '09':
                month = 'Sep'
            elif datestd[5:7] == '10':
                month = 'Oct'
            elif datestd[5:7] == '11':
                month = 'Nov'
            elif datestd[5:7] == '12':
                month = 'Dec'
        
            plastic_issue_date = (datestd[8:10]+'-'+month+'-'+datestd[0:4])
            return(plastic_issue_date)
        
        def ip_weight(cardaction):#=============================================ip weight assign
            weight_ip = ''
            try:   
                with open('config/AU_Credit_Variant_Weight_Table.csv', 'r') as ip_we:
                    contents = ip_we.readlines()
                    for line in contents:
                        content = line.split(',')
                        try:
                            if (cardaction == content[0]):
                                weight_ip = content[1].rstrip()
                            
                        except:
                            weight_ip = ''
            except:
                print('Error in AU_Credit_Variant_Weight_Table.csv against card_action : '+cardaction)
                
            
            return weight_ip
                            
        def awb_assign(courier_2):#=============================================awb_assign
            global bd_num,dl_num,ip_num,awb
            if courier_2 == 'Bluedart':
                try:
                    with open('config/BD_AWB.txt', 'r') as bdfin:
                        bdawb1 = bdfin.readlines()
                        awb_check = (len(bdawb1)-(bd_num + 1))
                        # if awb_check!=0:
                        awb = bdawb1[bd_num].rstrip()
                        bd_num += 1    
                except:
                    print("BD Awb Finished, Remaining Count : "+str(len(bdawb1)))
                    program_closed()
                         
                        
                    
            elif courier_2 == 'Delhivery':
                try:
                    with open('config/DL_AWB.txt', 'r') as dlfin:
                        dlawb = dlfin.readlines()
                        awb_check = (len(dlawb)-(dl_num + 1))
                        # if awb_check!=0:
                        awb = dlawb[dl_num].rstrip()
                        dl_num += 1
                            
                except:
                    print("DTDC Awb Finished, Remaining Count : "+str(len(dlawb)))
                    program_closed()
                         
                   
            elif courier_2 == 'Indiapost':
                try:
                    
                    with open('config/IP_AWB.txt', 'r') as ipfin:
                        ipawb = ipfin.readlines()
                        awb_check = (len(ipawb)-(ip_num + 1))
                        # if awb_check!=0:
                        awb = ipawb[ip_num].rstrip()
                        ip_num += 1
                except:
                    print("IP Awb Finished , Remaining Count : "+str(len(ipawb)))
                    program_closed()
                   
            return(awb)   
        
#=============================================Excel-Prcocessing starts=============================================           
        def excel_convertor(name_ff):
            #ext='.txt'
            if name_ff.__contains__('Zenith_Plus') or name_ff.__contains__('AU_Traverse'):
                if Path(name_ff).is_file():
                    #print('file present')
                    records_list = []
                    s_no = 0
                    with open(name_ff, "r") as ff, open(name_ff[2:], "w") as foutfile:
                        foutfile.write("Sr. No|Account Number|Primary Card number|Customer Name|Add-on Card 1|Add-on Card 2|Add-on Card 3|Add-on Card 4|PP  Card No_Primary|PP  Card No_Addon1|PP  Card No_Addon2|PP  Card No_Addon3|Ref no.|AWB. No.|Courier|Courier Code|Card Action|Primary Count|Add-on Count|Total Cards|Bin|Logo|Gender Code|Varient|Address Line 1|Address Line 2|Address Line 3|Address Line 4|City|State Code|Postal Code|Mobile Number|Cust Unique ID|Credit Limit|Statement Date |Extention|Embo_Name|Joining Fees|Annual Fees|Card Issuance Date|PP_Customer_Name|PP_Expire_Date|Card Type|Expiry Date|Product_code"+'\n')
                        contents = ff.readlines()
                        # sorting using our custom logic
                        contents.sort(key=my_sort_ff_courier)
                        contents.sort(key=my_sort_ff_cardtype)
                        
                        for fi in contents:
                            records_list.append(fi)
                            s_no = (records_list.index(fi)+1)
                            s_no = "%05d" % s_no
                            if not fi.strip():
                                continue
                            if fi:
                                foutfile.write(str(s_no)+fi)
                    ff.close()
                    foutfile.close()
                    df = pd.read_csv(name_ff[2:],encoding= 'unicode_escape', sep='|', dtype=object)
                    df.to_excel(name_ff[2:-4]+'_'+ptime[0:10]+'--'+ str(s_no)+'.xlsx', 'Sheet1', index=False)
            else :
                if Path(name_ff).is_file():
                    #print('file present')
                    records_list = []
                    s_no = 0
                    with open(name_ff, "r") as ff, open(name_ff[2:], "w") as foutfile:
                        foutfile.write("Sr. No|Account Number|Primary Card number|Customer Name|Add-on Card 1|Add-on Card 2|Add-on Card 3|Add-on Card 4|PP  Card No_Primary|PP  Card No_Addon1|PP  Card No_Addon2|PP  Card No_Addon3|Ref no.|AWB. No.|Courier|Courier Code|Card Action|Primary Count|Add-on Count|Total Cards|Bin|Logo|Gender Code|Varient|Address Line 1|Address Line 2|Address Line 3|Address Line 4|City|State Code|Postal Code|Mobile Number|Cust Unique ID|Credit Limit|Statement Date |Extention|Embo_Name|Joining Fees|Annual Fees|Card Issuance Date|Card Type|Expiry Date|Product_code"+'\n')
                        contents = ff.readlines()
                        # sorting using our custom logic
                        contents.sort(key=my_sort_ff_courier)
                        contents.sort(key=my_sort_ff_cardtype)
                        
                        for fi in contents:
                            records_list.append(fi)
                            s_no = (records_list.index(fi)+1)
                            s_no = "%05d" % s_no
                            if not fi.strip():
                                continue
                            if fi:
                                foutfile.write(str(s_no)+fi)
                    ff.close()
                    foutfile.close()
                    df = pd.read_csv(name_ff[2:],encoding= 'unicode_escape', sep='|', dtype=object)
                    df.to_excel(name_ff[2:-4]+'_'+ptime[0:10]+'--'+ str(s_no)+'.xlsx', 'Sheet1', index=False)
            os.remove(name_ff)
            os.remove(name_ff[2:])
                
        def excel_convertor_replace(name_ff):
            
            #ext='.txt'
            if name_ff.__contains__('Zenith_Plus') or name_ff.__contains__('AU_Traverse'):
                if Path(name_ff).is_file():
                    #print('file present')
                    records_list = []
                    s_no = 0
                    with open(name_ff, "r") as ff, open(name_ff[2:], "w") as foutfile:
                        foutfile.write("Sr. No|Account Number|Primary Card number|Customer Name|Add-on Card 1|Add-on Card 2|Add-on Card 3|Add-on Card 4|PP  Card No_Primary|PP  Card No_Addon1|PP  Card No_Addon2|PP  Card No_Addon3|Ref no.|AWB. No.|Courier|Courier Code|Card Action|Primary Count|Add-on Count|Total Cards|Bin|Logo|Gender Code|Varient|Address Line 1|Address Line 2|Address Line 3|Address Line 4|City|State Code|Postal Code|Mobile Number|Cust Unique ID|Credit Limit|Statement Date |Extention|Embo_Name|PP_Customer_Name|PP_Expire_Date|CardType|Expiry Date|Product_code"+'\n')
                        contents = ff.readlines()
                        # sorting using our custom logic
                        contents.sort(key=my_sort_ff_courier)
                        contents.sort(key=my_sort_ff_cardtype)
                        
                        for fi in contents:
                            records_list.append(fi)
                            s_no = (records_list.index(fi)+1)
                            s_no = "%05d" % s_no
                            if not fi.strip():
                                continue
                            if fi:
                                foutfile.write(str(s_no)+fi)
                    ff.close()
                    foutfile.close()
                    df = pd.read_csv(name_ff[2:],encoding= 'unicode_escape', sep='|', dtype=object)
                    df.to_excel(name_ff[2:-4]+'_'+ptime[0:10]+ '--'+str(s_no) +'.xlsx', 'Sheet1', index=False) 
                    
            else :
                
                if Path(name_ff).is_file():
                    #print('file present')
                    records_list = []
                    s_no = 0
                    with open(name_ff, "r") as ff, open(name_ff[2:], "w") as foutfile:
                        foutfile.write("Sr. No|Account Number|Primary Card number|Customer Name|Add-on Card 1|Add-on Card 2|Add-on Card 3|Add-on Card 4|PP  Card No_Primary|PP  Card No_Addon1|PP  Card No_Addon2|PP  Card No_Addon3|Ref no.|AWB. No.|Courier|Courier Code|Card Action|Primary Count|Add-on Count|Total Cards|Bin|Logo|Gender Code|Varient|Address Line 1|Address Line 2|Address Line 3|Address Line 4|City|State Code|Postal Code|Mobile Number|Cust Unique ID|Credit Limit|Statement Date |Extention|Embo_Name|CardType|Expiry Date|Product_code"+'\n')
                        contents = ff.readlines()
                        # sorting using our custom logic
                        contents.sort(key=my_sort_ff_courier)
                        contents.sort(key=my_sort_ff_cardtype)
                        
                        for fi in contents:
                            records_list.append(fi)
                            s_no = (records_list.index(fi)+1)
                            s_no = "%05d" % s_no
                            if not fi.strip():
                                continue
                            if fi:
                                foutfile.write(str(s_no)+fi)
                    ff.close()
                    foutfile.close()
                    df = pd.read_csv(name_ff[2:],encoding= 'unicode_escape', sep='|', dtype=object)
                    df.to_excel(name_ff[2:-4]+'_'+ptime[0:10]+ '--'+str(s_no) +'.xlsx', 'Sheet1', index=False)
                
            print(name_ff+'\n'+name_ff[2:])  
            os.remove(name_ff)
            os.remove(name_ff[2:])
                
        
        def embo_convertor(name_embo):
            #ext='.txt'
            if Path(name_embo).is_file():
                with open(name_embo, "r") as bf, open(name_embo[2:-4]+'_'+str(ptime)+'.txtt', "w") as boutfile:
                    contents = bf.readlines()
                    # sorting using our custom logic
                    contents.sort(key=my_sort_emb_courier)
                    contents.sort(key=my_sort_emb_cardtype)
                    
                    for bi in contents:
                        if not bi.strip():
                            continue
                        if bi:
                            boutfile.write(bi)
                            #total_primary_count+=1
                bf.close()
                boutfile.close()
            os.remove(name_embo)
            
        def pp_embo_convertor(pp_name_embo):
            if Path(pp_name_embo).is_file():  
                with open(pp_name_embo, "r") as bf, open(pp_name_embo[2:-4]+str(ptime)+'.txtt', "w") as boutfile:
                        contents = bf.readlines()
                        # sorting using our custom logic
                        # (key=pp_sort_emb)
                        contents.sort(key=pp_sort_emb_courier)
                        contents.sort(key=pp_sort_emb_cardtype)
                        for bi in contents:
                            if not bi.strip():
                                continue
                            if bi:
                                boutfile.write(bi)
                                #total_pp_count+=1
                bf.close()
                boutfile.close()
            os.remove(pp_name_embo)
            
            
        def pp_excel_convertor(pp_name_ff):
            #ext='.txt'
            if Path(pp_name_ff).is_file():
                #print('file present')
                records_list = []
                s_no = 0  
                with open(pp_name_ff, "r") as ff, open(pp_name_ff[2:], "w") as foutfile:
                    foutfile.write('Sr. No|Account Number|Ref no.|Primary Card number|Customer Name|PP  Card No_Primary|Expiry Date|Variant|Logo|Gender|AWB. No.|Courier|CardType'+'\n')
                    contents = ff.readlines()
                    # sorting using our custom logic
                    contents.sort(key=my_sort_pp_courier)
                    contents.sort(key=my_sort_pp_cardtype)
                    
                    for fi in contents:
                        records_list.append(fi)
                        s_no = (records_list.index(fi)+1)
                        s_no = "%05d" % s_no
                        if not fi.strip():
                            continue
                        if fi:
                            foutfile.write(str(s_no)+fi)
                ff.close()
                foutfile.close()
                df = pd.read_csv(pp_name_ff[2:],encoding= 'unicode_escape', sep='|', dtype=object)
                df.to_excel(pp_name_ff[2:-4]+'_'+ptime[0:10]+ '--'+str(s_no) +'.xlsx', 'Sheet1', index=False) 
                
                os.remove(pp_name_ff)
                os.remove(pp_name_ff[2:])
                # os.remove(name_embo+ext)
        
        def prod_var(embo_bin,embo_logo,embo_plastic_id,embo_gender_code_acc):
            # print(embo_bin+'|'+embo_logo+'|'+embo_plastic_id+'|'+embo_gender_code_acc)
            with open('config/product_varient_list.txt','r') as data_file:
                contents = data_file.readlines()
                file_embo,file_logo,file_plastic_id,file_gender,file_product,file_varient =  '','','','','',''
                embo_product =  'New Product'
                embo_varient =  'New Varient'
                for line in contents[1:]:
                    line_data = line.split(',')
                    file_embo = line_data[0]
                    file_logo = line_data[1]
                    file_plastic_id = line_data[2].upper()
                    file_gender = line_data[3].upper()
                    file_product = line_data[4].strip()
                    file_varient = line_data[5].strip()
                    
                    # if file_logo != 306:
                        
                    if embo_bin == file_embo and embo_logo == file_logo:
                        if embo_plastic_id.strip()  == file_plastic_id and embo_gender_code_acc ==  file_gender:
                            embo_product = file_product
                            embo_varient = file_varient
                            break

                        
                        elif 'NO'  ==  file_plastic_id and  embo_gender_code_acc == file_gender:
                            embo_product = file_product
                            embo_varient = file_varient
                            break
                        elif embo_plastic_id.strip()   == file_plastic_id  and  ''  == file_gender:
                            embo_product = file_product
                            embo_varient = file_varient
                            break
                        
                        elif 'NO'  == file_plastic_id and '' == file_gender:
                            embo_product = file_product
                            embo_varient = file_varient
                            break
            return(embo_product+'|'+embo_varient)
        
        
        
        def contains_special_characters(input_string):
            for char in input_string:
                asscii_value = ord(char)
                if 65 <= int(asscii_value) <= 90 or 97 <= int(asscii_value) <= 122 or int(asscii_value) == 32:
                    a = True
                else:
                    a = False
                    break
            return a
        
        
        def courier_summary(file,dmy,courier_sum_total_count):
            df = pd.read_excel(file)
            df['Bluedart count'] = (df['Courier'] == 'Bluedart').astype(int)
            df['Delhivery count'] = (df['Courier'] == 'Delhivery').astype(int)
            df['IndiaPost count'] = (df['Courier'] == 'Indiapost').astype(int)
            df['PP Total Count'] = (df['Total PP Count'] == 1).astype(int)
            df['Total count'] = (df['Courier'] == 'Bluedart').astype(int) + (df['Courier'] == 'Delhivery').astype(int) + (df['Courier'] == 'Indiapost').astype(int)
            # Group by 'Logo' and 'Product Name', then sum the counts
            result = df.groupby(['Logo', 'Varient','Bin']).agg(
                {
                    'Bluedart count': 'sum',
                    'Delhivery count': 'sum',
                    'IndiaPost count': 'sum',
                    'Total count': 'sum',
                    'PP Total Count': 'sum'
                }).reset_index()

            result.to_excel('AUF_CREDIT_COURIER_SUMMARY_FILE_'+dmy+'_'+str(courier_sum_total_count)+'.xlsx', index=False)

        def protect_excel(file_path, password):
            excel = None
            try:
                excel = win32.Dispatch('Excel.Application')
                excel.DisplayAlerts = False  # Suppress alerts like the "replace" dialog
                excel.Visible = False
                workbook = excel.Workbooks.Open(file_path)
                
                # Replace the extension from .xlsx to .xls
                new_file_path = file_path.replace('.xlsx', '.xls')
                
                # Save the file with the new .xls extension and overwrite the original
                workbook.SaveAs(new_file_path, 
                                Password=password,  
                                FileFormat=56)  # 56 is the format for .xls (Excel 97-2003)
                # print(f"File has been saved as .xls with password: {new_file_path}")
                
            except Exception as e:
                print(f"Error setting password for {file_path}: {e}")
            finally:
                if workbook:
                    workbook.Close(SaveChanges=True)
                if excel:
                    excel.DisplayAlerts = True  # Re-enable alerts after the operation
                    excel.Quit()
                    
                    
        def product_matrix(Product_name,card_type):
            card_type_code = ''
            with open('config/AUFC_PRODUCT_MATRIX.txt','r') as pd_file:
                contents = pd_file.readlines()
                if card_type == 'New Issue':
                    card_type_code = 'NW'
                elif card_type == 'Additional Card':
                    card_type_code = 'AD'
                elif card_type == 'Replacement Card' or card_type == 'Return Card':
                    card_type_code = 'DR'
                elif card_type == 'Emergency Card Requested':
                    card_type_code = 'ER'
                elif card_type == 'Reissue Card' or card_type == 'Reissue Card with different card numbering scheme' or card_type == 'Card technology reissue':
                    card_type_code = 'LS'
                elif card_type == 'Lost and Stolen Replacement':
                    # card_type_code = 'RW'
                    card_type_code = 'LS'
                
                for lin in contents[1:]:
                    val = lin.split(',')
                    Product_Name_matrix = val[0].strip()
                    if Product_name == Product_Name_matrix:
                        CODE         = val[1].strip()
                        return CODE+card_type_code
            
            return ''
        
        
        
        
        
#=========================================================================================           
#=============================================File-Prcocessing starts=============================================   
   
   
   
        file_name = glob.glob('L2CDZU*')
        print(file_name)
        if len(file_name)==0:
            print("No Input Files present")
        else:
            #=============================================File Dupe Check
            with open("config/File_DupeCheck.csv",'r') as dupecheck_file,open(f"AUF_File_DupeCheck_{ptime_1}.csv",'a') as dupecheck_file_cust:
                
                
                dupecheck_file_filecontents = dupecheck_file.readlines()
                Types = [line.split(":") for line in dupecheck_file_filecontents]
                filenames_list = [Type[1].rstrip() for Type in Types]
                # print(filenames_list)
                # print(file_name)
                
                #filenames = glob.glob('L2CDZU*')
                merge_output_file = 'del_output_file_'+str(x)+'_.txt'
                 
                for i in range(len(file_name)):
                    # print(file_name[file_count_num])
                    if file_name[file_count_num] in filenames_list:
                        pos_1 = filenames_list.index(file_name[file_count_num])
                        date_1_old_file = dupecheck_file_filecontents[pos_1][0:5]
                        # print(date_1_old_file)
                        date_1_old_file = "20" + date_1_old_file
                        # print(date_1_old_file)
                        # datestd = datetime.datetime.strptime(str(date_1_old_file), '%Y%j').date()
                        # print(datestd)
                        dupecheck_file_cust.write("Duplicate files found : "+str(file_name[file_count_num])+" old file date : "+jdtodatestd(date_1_old_file)+"\n")
                        
                        del file_name[file_count_num]
                        file_count_num-=1
                    else:
                        with open((file_name[file_count_num]), 'r') as filehandle:
                            filecontents = filehandle.readlines()
                            del filecontents[0]
                            del filecontents[-1]
                            
                            with open(merge_output_file, 'a') as outfile:
                                for fname in filecontents:
                                    outfile.write(fname)
                        filehandle.close()
                     
                    file_count_num+=1
           #=============================================  
          
            if len(file_name)==0:
                print("\n\nAll Input Files are duplciate within 15 days!!!")
                program_closed()
            
            #==========================================================================================
            merge_filename = glob.glob('del_output_file*.txt')
            with open('merge_output_file_'+str(x)+'.txt', 'w') as outfile:
                for fname in merge_filename:
                    with open(fname) as infile:
                        for line in infile:
                            outfile.write(line)
                            records_list.append(line)
            
                            # for i in enumerate(line):
                            account_number = line[38:54]
                            card_holder_type = line[2089:2090]
                            
                            if  account_number not in pri_account_number and  card_holder_type == "1":
                                pri_account_number.append(account_number)
                            elif  card_holder_type == "0":
                                add_on_account_number.append(account_number)
                            file_credit_card_em = line[:]
                                
            
                            #with open('AUC_EM_account_'+str(account_number)+'_'+str(ptime)+'.txt', 'a') as em_file:
                            with open('AUC_EM_account.txt', 'a') as em_file:
                                # em_file.write(file_credit_card_em)
                                em_file.write(embo_name_checker(file_credit_card_em))   #Replacing Embo with dots to spaces requested by Aamir sir
                            em_file.close()
                            
            outfile.close()
            
            for acc in add_on_account_number:
                if acc not in pri_account_number:
                    for i in range(int(add_on_account_number.count(acc))-1):
                        add_on_account_number.remove(acc)
            
            pri_account_number+=add_on_account_number #for toal no of cards as per 
                    
            #==========================================================================================
            auf_pp_output_header = 'Account_Number,Logo,Card_Number,Card_Holder_name,Card_action,Priority_Pass_Number,PP_issuance_date'
            
            with open('Priority_Pass_Data_'+str(ptime[0:10])+'.csv', 'w') as pp_file_auf_h:
                pp_file_auf_h.write(auf_pp_output_header + '\n')
                
                #print (file_pp_merge)
            pp_file_auf_h.close()
            
            with open('config\RTO_ADDRESS.txt', 'r') as IXIGOin:
                contents = IXIGOin.readlines()
                for line in contents:
                    data = line.split('|')
                    rto_address = data[0].rstrip()
                    rto_pin = data[1].rstrip()
            

            logo_gen_code=[]
            with open('config\Gender_Logo_code.csv', 'r') as gencode:
                contents = gencode.readlines()
                for ln in contents:
                   logo_gen_code.append(ln.replace('\n','')) 
            # print(logo_gen_code)
                    
            with open('config\Exclude_logo_retail_card_type.csv', 'r') as IXIGOin:
                contents = IXIGOin.readline()
                logos = contents.split(',')
                for logo in logos:
                    fourth_line_logo_list.append(logo)
                # print(fourth_line_logo_list)   
                    
            merge_filename_acc = glob.glob('*account*')
            with open('M_account_file_'+str(x)+'.txt', 'w') as outfile_acc:
                s_no1=0
                for fname_acc in merge_filename_acc:
                    with open(fname_acc) as infile_acc:
                        for line_acc in infile_acc:
                            outfile_acc.write(line_acc)
                            line_acc=line_acc.replace('â€™', ' ')
                            line_acc=line_acc.replace('|', ' ')
                            line_acc=line_acc.replace('â€œ', ' ')
                            line_acc=line_acc.replace('”', ' ')
                            line_acc=line_acc.replace('“', ' ')
                            line_acc=line_acc.replace('"', ' ')
                            line_acc=line_acc.replace('…', ' ')
                            # records_list_acc.append(line_acc)
                            s_no1 += 1 #(records_list.index(line_acc.rstrip()))
                            s_no = "%03d" % s_no1
                            account_number_acc = line_acc[38:54]
                            print_embo_name = line_acc[54:82]
                            
                            name_check = ''
                            name_check = contains_special_characters(print_embo_name)
                            if name_check == False:
                                with open("Special_Character_Records.csv",'a') as special_chr_rec:
                                    special_chr_rec.write(print_embo_name +' has special character in EMbo name acoount number details is '+account_number_acc+' in file name : '+'\n') 
                            else:
                                # for i in enumerate(line_acc):
                                logo_acc = line_acc[3:6]
                                # print(logo_acc)
                                c_bin_f_acc = line_acc[9:15] 
                                account_number_acc = line_acc[38:54]
                                # account_number_list.append(account_number_acc)
                                print_card_holder_name_acc = line_acc[54:80].rstrip()
                                # print(print_card_holder_name_acc)
                                pct_id = line_acc[501:504]
                                card_holder_name_acc = line_acc[1518:1558]
                                date_acc = line_acc[2659:2666]
                                card_holder_firstname_acc = line_acc[2763:2803].rstrip()
                                card_holder_middlename_acc = line_acc[2803:2843].rstrip()
                                card_holder_lastname_acc = line_acc[2843:2883].rstrip()
                                
                                name_of_customer = card_holder_firstname_acc + card_holder_middlename_acc + card_holder_lastname_acc
                                
                                plastic_id = line_acc[402:412].strip()
                                #salutation_acc = line_acc[2090:2094].rstrip()
                                card_holder_type_acc = line_acc[2089:2090]
                                cardnumber_acc = line_acc[9:25]
                                mask_pan_acc = cardnumber_acc[0:6] +'XXXXXX'+cardnumber_acc[12:16]
                                c_bin_acc = line_acc[9:15]
                                f_bin_acc = line_acc[9:15]
                                pp1_acc = line_acc[18:25]
                                pp2_reverse_acc = pp1_acc[::-1]
                                pp3_rand_acc = random.choices(pp2_reverse_acc, k=7)
                                pp4_acc = ''.join(map(str, pp3_rand_acc))
            
                                
            
                                print_exp_date_acc = line_acc[162:164] +'/'+line_acc[164:166]
                                exp_date_acc_ff = line_acc[162:164] +'/20'+line_acc[164:166]#ADDED ON 150524 AS PER youtrack request BANKING 108
                                pp_exp_date_acc = ptime[3:5]+ptime[8:10] #ADDED ON 23052022 AS PER CUSTOMER REQUIREMENT FOR CARD EXPIRY OF 3 YEARS
                                gender_code_acc = line_acc[2103:2104]
                                customer_id_acc1 = line_acc[1134:1153].lstrip()
                                customer_id_acc = customer_id_acc1.rstrip()
                                cr_card_limit = line_acc[167:182]
                                cr_card_limit = cr_card_limit.lstrip('0')
                                cr_card_limit = format_currency(cr_card_limit, '', locale='en_IN')
                                #cr_card_limit = format_currency(cr_card_limit, 'INR', locale='en_IN') #"INR" for indian rupee logo
                                #cr_card_limit=str(cr_card_limit)+'.00'
                                card_action_code = line_acc[150:151]
                                billing_cycle = line_acc[3844:3846]
                                billing_cycle_int = int(billing_cycle)
                                add_on_mask_pan = ''
                                tile_acc = line_acc[1938:1958].rstrip()
                                mailer_name1_acc = line_acc[1518:1558].rstrip()
                                mailer_name2_acc = line_acc[1558:1598].rstrip()
                                mailer_name3_acc = line_acc[1598:1638].rstrip()
                                
                                # mailer_name_acc = mailer_name1_acc+' '+mailer_name2_acc+' '+mailer_name3_acc
                                mailer_name_acc = mailer_name1_acc
                                mailer_address1_acc = line_acc[1638:1678].rstrip()
                                mailer_address2_acc = line_acc[1678:1718].rstrip()
                                mailer_address3_acc = line_acc[1718:1758].rstrip()
                                mailer_address4_acc = line_acc[1878:1918].rstrip()
                                mailer_city_acc = line_acc[1758:1788].rstrip()
                                mailer_state_code_acc = line_acc[1838:1868].rstrip()
                                mailer_postal_code_acc = line_acc[2016:2026].rstrip()
                                mailer_mob_acc = line_acc[3143:3163].rstrip() 
                                suffix = ""
                                pp_cardnumber_acc = "'NA"
                                pp_count=''
                                totol_card_count = pri_account_number.count(account_number_acc)
                                courier = courier_assign(mailer_postal_code_acc)[0]
                                courier = courier.title()
                                routing_code = courier_assign(mailer_postal_code_acc)[1]
                                
                                
                                # name_check = contains_special_characters(name_of_customer)
                                
                                
                                
                                    
        # =============================================================================
        #                         if card_holder_type_acc == '1' or (card_holder_type_acc == '0' and account_number_acc not in account_number_list):
        #                             # row=[]
        #                             #card_type = 'Primary Card'
        #                             # account_number_list.append(account_number_acc)
        #                             
        #                             awb_number = awb_assign(courier)
        #                             
        #         
        #                         elif card_holder_type_acc == '0' and account_number_acc in account_number_list:
        #                             #index=np.where(arr == 15)
        #                             pos = account_number_list.index(account_number_acc)
        #                             awb_number = account_number_list[pos+1]
        #                             #add_on_mask_pan = mask_pan_acc
        #                         else:
        #                             awb_number = ''
        #                             #add_on_mask_pan = ''
        #                         # if card_holder_type_acc == '1':
        #                         
        #                         if card_holder_type_acc == '1':
        #                             account_number_list.extend([account_number_acc, awb_number])
        # =============================================================================
                                # if name_check == False:
                                #     with open('Incorrect_name_Embo_file.txt','a') as error_file:
                                #         error_file.write(cardnumber_acc+'|'+name_of_customer+'\n')
                                        
                                # else:
                                
                                
                                
                                
                                
                                
                                
                                if gender_code_acc == '1':
                                    salutation_acc = 'Mr.'
                                elif gender_code_acc == '2':
                                    salutation_acc = 'Ms.'
                                else:
                                    salutation_acc = 'Not Found'
            
                                
                                if 4 <= billing_cycle_int <= 20 or 24 <= billing_cycle_int <= 30:
                                    suffix = "th"
                                else:
                                    suffix = ["st", "nd", "rd"][billing_cycle_int % 10 - 1]
                                superscript = suffix
            
                                
            
                                if plastic_id == '0000000002':
                                    celeb = 'C_Male'
                                elif plastic_id == '0000000003':
                                    celeb = 'C_Female'
                                else:
                                    celeb = ''
                                
                                   
                                    
                                j_ts = date.today()
                                j_x = (j_ts)
                                j_fmt = '%Y-%m-%d'
                                j_s = str(j_x)
                                j_dt = datetime.datetime.strptime(j_s, j_fmt)
                                j_tt = j_dt.timetuple()
                                j_tt.tm_yday
                                j_p = int('%d%03d' % (j_tt.tm_year, j_tt.tm_yday))
                                j_day = str(j_p)
            
                                ref_number = customer_id_acc[2:]+card_holder_name_acc[0:4].upper()+j_day[4:7]+j_day[0:4]+str(s_no)
                                if card_holder_type_acc == '1':
                                    card_type = 'Primary Card'
                                    # account_number_list.append(account_number_acc)
                                elif card_holder_type_acc == '0':
                                    card_type = 'Add-on  Card'
            
                                ps = account_number_acc+'|'+card_type
                                pps = [account_number_acc, mask_pan_acc]
                                allocation_gender_code = gender_code_acc
                                if logo_acc not in logo_gen_code:
                                    gender_code_acc = ''


                                # if logo_acc == '101' and c_bin_acc == '466505' and gender_code_acc == '1' and plastic_id = '':
                                #     input()

                                var_prod = prod_var(str(c_bin_acc),str(logo_acc),str(plastic_id),str(gender_code_acc))
                                var_prod = var_prod.split('|')
                                product = var_prod[0]
                                varient = var_prod[1]
                                
                                

                                varient = varient.replace('AU_BANK_LIT_CREDIT_CARD','LIT')
                                
            
                                if card_action_code == '1':
                                    card_action = 'New Issue'  
                                elif card_action_code == '2':
                                    card_action = 'Additional Card'
                                elif card_action_code == '3':
                                    card_action = 'Replacement Card'
                                elif card_action_code == '4':
                                    card_action = 'Return Card'
                                elif card_action_code == '6':
                                    card_action = 'Emergency Card Requested'
                                elif card_action_code == '7':
                                    card_action = 'Reissue Card'
                                elif card_action_code == '8':
                                    card_action = 'Reissue Card with different card numbering scheme'
                                elif card_action_code == '9':
                                    card_action = 'Card technology reissue'
                                elif card_action_code == 'A':
                                    card_action = 'Additional Card'
                                elif card_action_code == 'L':
                                    card_action = 'Lost and Stolen Replacement'
                                elif card_action_code == '0':
                                    card_action = 'No Action Requested'
                                else:
                                    card_action = 'No action code matched'
                                
                                if card_action[0:10] == "Additional":
                                    card_action_1 = "Add-On"
                                else:
                                    card_action_1 = "Pri"
                                    
                                    
                                ip_weigh = ip_weight(varient)
                                if ip_weigh == '':
                                    ip_weigh = ip_weight(card_action)
                                    
                                
                                
                                
                                #print(card_holder_name_acc)
                                joining_fees = fees(logo_acc,pct_id)[0]
                                annual_fees = fees(logo_acc,pct_id)[1]
                                
                                product_code = product_matrix(product,card_action)       
                                
                                # Assigning awb on cards
                                try: 
                                    if totol_card_count == 1 or product.__contains__("Zenith_Plus") or product.__contains__("LIT") or product.__contains__("AU_KOSMO"):
                                    # if totol_card_count == 1 :
                                        awb_number = awb_assign(courier)
                                        # adding for total card count for Zenith_Plus or LIT
                                        totol_card_count = 1
                                    
                                    elif account_number_acc not in account_number_list:
                                            #index=np.where(arr == 15)
                                            awb_number = awb_assign(courier)
                                            account_number_list.extend([account_number_acc, awb_number])
                                    
                                    elif account_number_acc in account_number_list:        
                                            pos = account_number_list.index(account_number_acc)
                                            awb_number = account_number_list[pos+1]
                                            #add_on_mask_pan = mask_pan_acc
                                    else:
                                        awb_number = ''
                                except Exception as e:
                                    print("Error while assigning awb number : error msg : "+ e)
                                
                                    
                                
                                
                                
                                
                                
                                    
                                if card_holder_type_acc=='0':
                                    totol_addon_card_count = '1'
                                else:
                                    totol_addon_card_count = '0'
                                totol_card_count = str(totol_card_count)
                                retail_card_type = '                 '
                                # fourth_line_logo_list = ['803','911','101','701','702','703','601','602','603']
                                
                                if plastic_id == '0000000015' and logo_acc not in fourth_line_logo_list:
                                    retail_card_type = 'FIELD CLUB MEMBER'
                                    with open('FIELD_CLUB_MEMBER_FILE.txt','a') as fc_file:
                                        fc_file.write(mask_pan_acc+'|'+print_card_holder_name_acc+'|'+logo_acc+'|'+varient+'|'+retail_card_type+'|'+plastic_id)
                                    fc_file.close()

                                varient = varient+celeb

                                file_credit_card_ff_acc = ' '+'|'+c_bin_f_acc+'|'+logo_acc+'|'+gender_code_acc+'|'+varient+'|'+ref_number+'|'+card_type+'|'+mask_pan_acc+'|'+account_number_acc+'|'+card_action+'|'+salutation_acc+'|'+mailer_name_acc+'|'+mailer_address1_acc+'|'+mailer_address2_acc+'|' + mailer_address3_acc+'|'+mailer_address4_acc+'|'+mailer_city_acc+'|'+mailer_state_code_acc+'|'+mailer_postal_code_acc+'|'+mailer_mob_acc+'|' + customer_id_acc+'|'+cr_card_limit+'|'+billing_cycle+'|'+superscript+'|' + print_card_holder_name_acc+'|' +  pct_id+'|'  # +jdtodatestd(date_acc)
                                file_credit_card_ff_acc_allocation = ' '+'|'+account_number_acc + '|'+mask_pan_acc+'|'+mailer_name_acc+'|'+'|'+'|'+'|'+'|'
                                file_credit_card_ff_acc_primary = ' '+'|'+account_number_acc + '|'+mask_pan_acc+'|'+mailer_name_acc+'|'+'|'+'|'+'|'+'|'
                                file_credit_card_ff_acc_addon = ' '+'|'+account_number_acc + '|'+'|'+mailer_name_acc+'|'+mask_pan_acc+'|'+'|'+'|'+'|'
                                file_credit_card_ff_acc_3_primary = card_action+'|'+'1|'+totol_addon_card_count+'|'+totol_card_count+'|'+c_bin_f_acc+'|'+logo_acc+'|'+gender_code_acc+'|'+varient+'|'+mailer_address1_acc+'|'+mailer_address2_acc+'|'+mailer_address3_acc+'|'+mailer_address4_acc+'|'+mailer_city_acc + '|'+mailer_state_code_acc+'|'+mailer_postal_code_acc+'|'+mailer_mob_acc+'|'+customer_id_acc+'|'+cr_card_limit+'|' + billing_cycle+'|'+superscript+'|'+print_card_holder_name_acc + '|'+joining_fees+'|'+annual_fees + '|'+str(jdtodatestd(date_acc))
                                file_credit_card_ff_acc_3_primary_replace = card_action+'|'+'1|'+totol_addon_card_count+'|'+totol_card_count+'|'+c_bin_f_acc+'|'+logo_acc+'|'+gender_code_acc+'|'+varient+'|'+mailer_address1_acc+'|'+mailer_address2_acc+'|'+mailer_address3_acc+'|'+mailer_address4_acc+'|'+mailer_city_acc + '|'+mailer_state_code_acc+'|'+mailer_postal_code_acc+'|'+mailer_mob_acc+'|'+customer_id_acc+'|'+cr_card_limit+'|' + billing_cycle+'|'+superscript+'|'+print_card_holder_name_acc
                                file_credit_card_ff_acc_3_addon = card_action+'|'+'0|'+totol_addon_card_count+'|'+totol_card_count+'|'+c_bin_f_acc+'|'+logo_acc+'|'+gender_code_acc+'|'+varient+'|'+mailer_address1_acc+'|'+mailer_address2_acc+'|'+mailer_address3_acc+'|'+mailer_address4_acc+'|'+mailer_city_acc +'|'+mailer_state_code_acc+'|'+mailer_postal_code_acc+'|'+mailer_mob_acc+'|'+customer_id_acc+'|'+cr_card_limit+'|' +billing_cycle+'|'+superscript+'|'+print_card_holder_name_acc +'|'+joining_fees+'|'+annual_fees +'|'+str(jdtodatestd(date_acc))
                                file_credit_card_ff_acc_3_addon_replace = card_action+'|'+'0|'+totol_addon_card_count+'|'+totol_card_count+'|'+c_bin_f_acc+'|'+logo_acc+'|'+gender_code_acc+'|'+varient+'|'+mailer_address1_acc+'|'+mailer_address2_acc+'|'+mailer_address3_acc+'|'+mailer_address4_acc+'|'+mailer_city_acc +'|'+mailer_state_code_acc+'|'+mailer_postal_code_acc+'|'+mailer_mob_acc+'|'+customer_id_acc+'|'+cr_card_limit+'|' +billing_cycle+'|'+superscript+'|'+print_card_holder_name_acc 
                                indianpost_courier_1 = '||'+ref_number+'|'+mailer_city_acc+'|'+mailer_postal_code_acc+'|'+mailer_name_acc+'|' + mailer_address1_acc+'|'+mailer_address2_acc+'|'+mailer_address3_acc + '|'+mailer_address4_acc+'|'+mailer_mob_acc+'||'
                                file_credit_card_ff_acc_new = ' '+'|'+account_number_acc+'|'+ref_number+'|'+mask_pan_acc+'|'+salutation_acc+' '+mailer_name_acc + '|'+pp_cardnumber_acc+'|'+print_exp_date_acc+'|'+varient + '|'+logo_acc+'|' + gender_code_acc  # +jdtodatestd(date_acc)
                                file_credit_card_em_acc = line_acc[:].rstrip('\n')+'|'+retail_card_type
                                pct_id_data = str(ptime[0:10])+','+account_number_acc+','+logo_acc+','+pct_id+','+mask_pan_acc+','+card_holder_name_acc.rstrip()+','+card_action
                                #header='Sr. No|Account Number|Primary Card number|Customer Name|Add-on Card 1|Add-on Card 2|Add-on Card 3|Add-on Card 4|PP  Card No_Primary|PP  Card No_Addon1|PP  Card No_Addon2|PP  Card No_Addon3|Ref no.|AWB. No.|Courier|Courier Code|Card Action|Primary Count|Add-on Count|Total Cards|Bin|Logo|Gender Code|Varient|Address Line 1|Address Line 2|Address Line 3|Address Line 4|City|State Code|Postal Code|Mobile Number|Cust Unique ID|Credit Limit|Statement Date |Extention|Embo_Name|Joining Fees|Annual Fees|Card Issuance Date'
                                bin_list.append(c_bin_acc+logo_acc+gender_code_acc)
                                bin_list = list(dict.fromkeys(bin_list))
                                qty = ''    
                
                                 
                                #ADDITONAL DETAIL FOR ADD CARDS
                                ps1 = account_number_acc+'|'+card_type+'|'+mask_pan_acc+'|'+card_action+'|'+c_bin_f_acc+'|'+logo_acc+'|' +gender_code_acc+'|'+varient+'|'+mailer_name_acc
                                
                                
                                with open('xxAUF_FF_MIS_MERGE_'+str(ptime)+'.txt', 'a') as ff_file_accm:
                                    date_q = jdtodatestd(date_acc)
                                    date_q = str(date_q)
                
                                    ff_file_accm.write(file_credit_card_ff_acc+str(awb_number)+'|'+routing_code+'|'+courier+'|'+date_q+'\n')
                
                                ff_file_accm.close()
                                
                
                                ps_list.append(ps)
                                ps1_list.append(ps1)
                                pps_list.append(pps)
                
                                cardnumber_list.append(cardnumber_acc)
                                if logo_acc != '306':
                                    if ((c_bin_acc == '465523_1' and card_holder_type_acc=='1') or (c_bin_acc == '457036')  or (c_bin_acc == '653024' and logo_acc == '910') or (c_bin_acc == '653023' and  logo_acc == '912')):#and logo_acc == '402'):
                                        if (card_action_code == '1' or card_action_code == '2' or card_action_code == '7' or card_action_code == 'A'):
                                            if card_holder_type_acc == '1':
                                                # pp_cardnumber_acc = pp_cardnumber_acc+'01'
                                                primary_card.append(cardnumber_acc)
                            
                                            if card_holder_type_acc == '0':
                                                secondary_card.append(cardnumber_acc)
                            
                                    # pp_card card awb allocation
                                   
                                    if ((c_bin_acc == '465523_1' and card_holder_type_acc=='1') or (c_bin_acc == '457036') or (c_bin_acc == '653024' and logo_acc == '910') or (c_bin_acc == '653023' and  logo_acc == '912') ):#and logo_acc == '402'):
                                        if (card_action_code == '1' or card_action_code == '2' or card_action_code == '7' or card_action_code == 'A'):
                                            if c_bin_acc == '465523_1' :#Vetta cards
                                                with open('config/VETTA_PP_NUMBER.csv', 'r') as vetta_pp_in:
                                                    try:
                                                        vetta_pp_total    = vetta_pp_in.readlines()
                                                        pp_cardnumber_acc = vetta_pp_total[vetta_pp].rstrip()
                                                        vetta_pp+=1
                                                    except:
                                                        print("VETTA PP NUMBER Finished!!!! Remaining Count : "+str(len(vetta_pp_total)))
                                                        program_closed()
                                                    
                                            
                                                
                                            # elif (c_bin_acc == '457036' and logo_acc == '402') or (c_bin_acc == '653023' and (logo_acc == '910' or logo_acc == '912')):#Zenith cards
                                            elif ((c_bin_acc == '457036') or (c_bin_acc == '653024' and logo_acc == '910') or (c_bin_acc == '653023' and  logo_acc == '912')):#Zenith cards
                                            
                                                try:
                                                    with open('config/ZENITH_PP_NUMBER.csv', 'r') as zenith_pp_in:
                                                        zenith_pp_total   = zenith_pp_in.readlines()
                                                        pp_cardnumber_acc = zenith_pp_total[zenith_pp].rstrip()
                                                        zenith_pp+=1
                                                except:
                                                    print("ZENITH PP NUMBER Finished!!!!   Remaining Count : "+str(len(zenith_pp_total)))
                                                    program_closed()
                                                
                                            # print(c_bin_acc+'|'+logo_acc+'|'+pp_cardnumber_acc)
                                            pp_cardnumber_acc = pp_dup_check(pp_cardnumber_acc)
                                            
                                            pp_count = '1'
                                
                               
                                track1_acc = '%PP/'+salutation_acc+'/'+print_card_holder_name_acc+'//''?'
                               
                                track2_acc = ';'+pp_cardnumber_acc+'=' +(pp_exp_date_acc[0:2]+'20'+str(int(pp_exp_date_acc[2:4])+3)) + '?'
                                
                                file_credit_card_ff_acc_2_primary = pp_cardnumber_acc+'|'+'|'+'|' +'|'+ref_number+'|'+str(awb_number) +'|'+courier+'|'+routing_code+'|'
                                
                                file_credit_card_ff_acc_2_addon = pp_cardnumber_acc+'|'+'|'+'|' '|'+ref_number+'|'+str(awb_number) + '|'+courier+'|'+routing_code+'|'
                
                                auf_pp_output = account_number_acc+','+logo_acc+','+mask_pan_acc+',' +card_holder_name_acc.rstrip()+','+card_action+',' +pp_cardnumber_acc+','+str(ptime[0:10])
                                
                                auf_pp_output_bank = pp_cardnumber_acc+','+str(ptime[0:10])+','+(pp_exp_date_acc[0:2]+'20'+str(int(pp_exp_date_acc[2:4])+3))+','+account_number_acc+','+card_holder_name_acc.rstrip()
                                
                                bluedart_courier = ref_number+'|'+mailer_name_acc+'|'+mailer_address1_acc+'|'+mailer_address2_acc+'|'+mailer_address3_acc+'-' + mailer_address4_acc+'|'+mailer_city_acc+'|'+mailer_state_code_acc +'||'+mailer_postal_code_acc+'|' +mailer_mob_acc+'|'+str(awb_number)
                                
                                delhivery_courier = str(awb_number)+'|'+ref_number+'|'+mailer_name_acc+'|'+mailer_city_acc+'|'+mailer_state_code_acc+'||'+mailer_address1_acc+' '+mailer_address2_acc+' '+mailer_address3_acc+' '+mailer_address4_acc+'|'+mailer_postal_code_acc+'|'+mailer_mob_acc+'|' + '250|Prepaid|500|Secure Deliverables|'+rto_address+'|'+rto_pin+'|AU Small Finance Bank Ltd|AU Small Finance Bank Limited AU Centre, 3rd, 5th, 6th & 7th Floor, Sunny Trade Centre, New Atish Market, Jaipur, Rajasthan  302019|True|True'
                                
                                indianpost_courier = '||'+ref_number+'|'+mailer_city_acc+'|'+mailer_postal_code_acc+'|'+mailer_name_acc+'|'+mailer_address1_acc +'|'+mailer_address2_acc+'|'+mailer_address3_acc+'|' +mailer_address4_acc+'|'+mailer_mob_acc+'||'+str(awb_number)  +'|'+str(ip_weigh)
                                
                                allocation_data_primary = '|'+account_number_acc+'|'+mask_pan_acc+'|'+print_card_holder_name_acc+'|||||'+pp_cardnumber_acc+'||||'+ref_number+'|' +str(awb_number)+'|'+courier+'|'+routing_code+'|'+card_action+'|1|'+totol_addon_card_count+'|'+totol_card_count+'|'+c_bin_f_acc+'|'+logo_acc+'|'+allocation_gender_code+'|'+varient+'|'+mailer_address1_acc+'|'+mailer_address2_acc+'|'+mailer_address3_acc+'|' +mailer_address4_acc+'|'+mailer_city_acc+'|'+mailer_state_code_acc+'|'+mailer_postal_code_acc+'|' +mailer_mob_acc+'|'+customer_id_acc+'|'+pp_count+'|'+joining_fees +'|'+annual_fees+'|'+str(jdtodatestd(date_acc)+'|'+retail_card_type+'|'+plastic_id)                
                                
                                scanning_data = allocation_data_primary.replace('\n','') +'|'+mask_pan_acc[-4:]+'\n'

                                allocation_data_addon = '|'+account_number_acc+'||'+card_holder_name_acc.rstrip()+'|'+mask_pan_acc+'|||||'+pp_cardnumber_acc+'|||'+ref_number+'|'+str(awb_number)+'|'+courier+'|'+routing_code+'|'+card_action+'-Additional Card'+'|0|'+totol_addon_card_count+'|'+totol_card_count+'|'+c_bin_f_acc+'|'+logo_acc+'|'+allocation_gender_code+'|'+varient+'|'+mailer_address1_acc +'|'+mailer_address2_acc+'|'+mailer_address3_acc+'|'+mailer_address4_acc+'|'+mailer_city_acc+'|'+mailer_state_code_acc+'|' +mailer_postal_code_acc+'|'+mailer_mob_acc+'|'+customer_id_acc+'|' +pp_count+'|'+joining_fees+'|'+annual_fees +    '|'+str(jdtodatestd(date_acc)+'|'+retail_card_type+'|'+plastic_id)
                                scanning_data = allocation_data_addon.replace('\n','') +'|'+mask_pan_acc[-4:]+'\n'
                                if logo_acc != '306':
                                    
                                    if ((c_bin_acc == '465523_1' and card_holder_type_acc=='1') or (c_bin_acc == '457036') or (c_bin_acc == '653024' and logo_acc == '910') or (c_bin_acc == '653023' and  logo_acc == '912')):#and logo_acc == '402'):
                                        if (card_action_code == '1' or card_action_code == '2' or card_action_code == '7' or card_action_code == 'A'):
                                            if mask_pan_acc not in filecontents:
                                                file_pp = '#0#'+mask_pan_acc+'#1#'+pp_cardnumber_acc+'#2#'+track1_acc+'#3#'+track2_acc +'#4#'+ (pp_exp_date_acc[0:2]+'/'+str(int(pp_exp_date_acc[2:4])+3)) +'#5#'+salutation_acc +' '+print_card_holder_name_acc+'#6#'
                                                
                                                file_pp_merge = account_number_acc+'|'+ref_number+'|'+mask_pan_acc+'|'+salutation_acc+' ' +print_card_holder_name_acc+'|'+pp_cardnumber_acc+'|' + (pp_exp_date_acc[0:2]+'/'+str(int(pp_exp_date_acc[2:4])+3)) +'|'+varient+'|'+logo_acc+'|'+gender_code_acc
                    
                                if courier == 'Indiapost':
                                    with open('xxxAUF_Credit_IndiaPost_Courier_Connection_Report_.txt', 'a') as ff_ip_acc:
                                        # ff_ip_acc.write(indianpost_courier_1+indianpost_courier_2+'\n')
                                        ff_ip_acc.write(indianpost_courier+'\n')
                                    ff_ip_acc.close()
                                elif courier == 'Delhivery':
                                    with open('xxxAUF_Credit_Delhivery_Courier_Connection_Report_.txt', 'a') as ff_dl_acc:
                                        ff_dl_acc.write(delhivery_courier+'\n')
                                    ff_dl_acc.close()
                                elif courier == 'Bluedart':
                                    with open('xxxAUF_Credit_Bluedart_Courier_Connection_Report_.txt', 'a') as ff_bd_acc:
                                        ff_bd_acc.write(bluedart_courier+'\n')
                                    ff_bd_acc.close()
                                
                                if product == 'Zenith_Plus' or product == 'Zenith_Plus_RuPay' or  product == 'AU_Traverse':
                                    if ((card_holder_type_acc == '1') and (card_action_code=='3' or card_action_code== 'L')):
                                        
                                        with open('xxAUF_Replace_FF_MIS_B'+str(c_bin_acc)+'_L'+str(logo_acc)+'_G'+str(gender_code_acc)+'_'+str(product).replace('AU_BANK_LIT_CREDIT_CARD','Lit_Credit_Card')+'_'+celeb+'.txt', 'a') as ff_file_acc:
                                            ff_file_acc.write(file_credit_card_ff_acc_primary +file_credit_card_ff_acc_2_primary+file_credit_card_ff_acc_3_primary_replace+'|'+salutation_acc+' ' +print_card_holder_name_acc+'|'+ (pp_exp_date_acc[0:2]+'/'+str(int(pp_exp_date_acc[2:4])+3)) +'|Pri_Replace|'+exp_date_acc_ff+'|'+product_code+'\n')
                                        ff_file_acc.close()
                                        
                                        with open('xxAUC_EM_B'+str(c_bin_acc)+'_L'+str(logo_acc)+'_G'+str(gender_code_acc)+'_'+str(product).replace('AU_BANK_LIT_CREDIT_CARD','Lit_Credit_Card')+'_'+celeb+'.txt', 'a') as em_file_acc:
                                            em_file_acc.write(file_credit_card_em_acc+'|'+courier+'|Pri_Replace'+'\n')
                                        em_file_acc.close()
                                        
                                    elif ((card_holder_type_acc == '0') and (card_action_code=='3' or card_action_code== 'L')):
                                        with open('xxAUF_Replace_FF_MIS_B'+str(c_bin_acc)+'_L'+str(logo_acc)+'_G'+str(gender_code_acc)+'_'+str(product).replace('AU_BANK_LIT_CREDIT_CARD','Lit_Credit_Card')+'_'+celeb+' - Add-On-Card.txt', 'a') as ff_addonfile_acc:
                                            ff_addonfile_acc.write(file_credit_card_ff_acc_addon+file_credit_card_ff_acc_2_addon+file_credit_card_ff_acc_3_addon_replace+'|'+salutation_acc+' ' +print_card_holder_name_acc+'|'+ (pp_exp_date_acc[0:2]+'/'+str(int(pp_exp_date_acc[2:4])+3)) +'|AddOn_Replace|'+exp_date_acc_ff+'|'+product_code+'\n')
                                        ff_addonfile_acc.close()
                                        
                                        with open('xxAUC_EM_B'+str(c_bin_acc)+'_L'+str(logo_acc)+'_G'+str(gender_code_acc)+'_'+str(product).replace('AU_BANK_LIT_CREDIT_CARD','Lit_Credit_Card')+'_'+celeb+'.txt', 'a') as em_file_acc:
                                            em_file_acc.write(file_credit_card_em_acc+'|'+courier+'|AddOn_Replace'+'\n')
                                        em_file_acc.close()
                                        
                                    elif card_holder_type_acc == '1' :
                                        with open('xxAUF_FF_MIS_B'+str(c_bin_acc)+'_L'+str(logo_acc)+'_G'+str(gender_code_acc)+'_'+str(product).replace('AU_BANK_LIT_CREDIT_CARD','Lit_Credit_Card')+'_'+celeb+'.txt', 'a') as ff_file_acc:
                                            ff_file_acc.write(file_credit_card_ff_acc_primary +file_credit_card_ff_acc_2_primary+file_credit_card_ff_acc_3_primary+'|'+salutation_acc+' ' +print_card_holder_name_acc+'|'+ (pp_exp_date_acc[0:2]+'/'+str(int(pp_exp_date_acc[2:4])+3)) +'|Pri|'+exp_date_acc_ff+'|'+product_code+'\n')
                                        ff_file_acc.close()
                                        
                                        with open('xxAUC_EM_B'+str(c_bin_acc)+'_L'+str(logo_acc)+'_G'+str(gender_code_acc)+'_'+str(product).replace('AU_BANK_LIT_CREDIT_CARD','Lit_Credit_Card')+'_'+celeb+'.txt', 'a') as em_file_acc:
                                            em_file_acc.write(file_credit_card_em_acc+'|'+courier+'|Pri'+'\n')
                                        em_file_acc.close()
                                        
                                    elif card_holder_type_acc == '0':
                                        with open('xxAUF_FF_MIS_B'+str(c_bin_acc)+'_L'+str(logo_acc)+'_G'+str(gender_code_acc)+'_'+str(product).replace('AU_BANK_LIT_CREDIT_CARD','Lit_Credit_Card')+'_'+celeb+' - Add-On-Card.txt'+'.txt', 'a') as ff_addonfile_acc:
                                            ff_addonfile_acc.write(file_credit_card_ff_acc_addon+file_credit_card_ff_acc_2_addon+file_credit_card_ff_acc_3_addon+'|'+salutation_acc+' ' +print_card_holder_name_acc+'|'+ (pp_exp_date_acc[0:2]+'/'+str(int(pp_exp_date_acc[2:4])+3)) +'|Add-On|'+exp_date_acc_ff+'|'+product_code+'\n')
                                        ff_addonfile_acc.close()
                                        
                                        with open('xxAUC_EM_B'+str(c_bin_acc)+'_L'+str(logo_acc)+'_G'+str(gender_code_acc)+'_'+str(product).replace('AU_BANK_LIT_CREDIT_CARD','Lit_Credit_Card')+'_'+celeb+'.txt', 'a') as em_file_acc:
                                            em_file_acc.write(file_credit_card_em_acc+'|'+courier+'|Add-On'+'\n')
                                        em_file_acc.close()
                                
                                else:
                                    if ((card_holder_type_acc == '1') and (card_action_code=='3' or card_action_code== 'L')):
                                        
                                        with open('xxAUF_Replace_FF_MIS_B'+str(c_bin_acc)+'_L'+str(logo_acc)+'_G'+str(gender_code_acc)+'_'+str(product).replace('AU_BANK_LIT_CREDIT_CARD','Lit_Credit_Card')+'_'+celeb+'.txt', 'a') as ff_file_acc:
                                            ff_file_acc.write(file_credit_card_ff_acc_primary +file_credit_card_ff_acc_2_primary+file_credit_card_ff_acc_3_primary_replace+'|Pri_Replace|'+exp_date_acc_ff+'|'+product_code+'\n')
                                        ff_file_acc.close()
                                        
                                        with open('xxAUC_EM_B'+str(c_bin_acc)+'_L'+str(logo_acc)+'_G'+str(gender_code_acc)+'_'+str(product).replace('AU_BANK_LIT_CREDIT_CARD','Lit_Credit_Card')+'_'+celeb+'.txt', 'a') as em_file_acc:
                                            em_file_acc.write(file_credit_card_em_acc+'|'+courier+'|Pri_Replace'+'\n')
                                        em_file_acc.close()
                                       
                                    elif ((card_holder_type_acc == '0') and (card_action_code=='3' or card_action_code== 'L')):
                                        
                                        with open('xxAUF_Replace_FF_MIS_B'+str(c_bin_acc)+'_L'+str(logo_acc)+'_G'+str(gender_code_acc)+'_'+str(product).replace('AU_BANK_LIT_CREDIT_CARD','Lit_Credit_Card')+'_'+celeb+' - Add-On-Card.txt', 'a') as ff_addonfile_acc:
                                            ff_addonfile_acc.write(file_credit_card_ff_acc_addon+file_credit_card_ff_acc_2_addon+file_credit_card_ff_acc_3_addon_replace+'|AddOn_Replace|'+exp_date_acc_ff+'|'+product_code+'\n')
                                        ff_addonfile_acc.close()
                                        
                                        with open('xxAUC_EM_B'+str(c_bin_acc)+'_L'+str(logo_acc)+'_G'+str(gender_code_acc)+'_'+str(product).replace('AU_BANK_LIT_CREDIT_CARD','Lit_Credit_Card')+'_'+celeb+'.txt', 'a') as em_file_acc:
                                            em_file_acc.write(file_credit_card_em_acc+'|'+courier+'|AddOn_Replace'+'\n')
                                        em_file_acc.close()
                                        
                                    elif card_holder_type_acc == '1' :
                                        
                                        with open('xxAUF_FF_MIS_B'+str(c_bin_acc)+'_L'+str(logo_acc)+'_G'+str(gender_code_acc)+'_'+str(product).replace('AU_BANK_LIT_CREDIT_CARD','Lit_Credit_Card')+'_'+celeb+'.txt', 'a') as ff_file_acc:
                                            ff_file_acc.write(file_credit_card_ff_acc_primary +file_credit_card_ff_acc_2_primary+file_credit_card_ff_acc_3_primary+'|Pri|'+exp_date_acc_ff+'|'+product_code+'\n')
                                        ff_file_acc.close()
                                        
                                        with open('xxAUC_EM_B'+str(c_bin_acc)+'_L'+str(logo_acc)+'_G'+str(gender_code_acc)+'_'+str(product).replace('AU_BANK_LIT_CREDIT_CARD','Lit_Credit_Card')+'_'+celeb+'.txt', 'a') as em_file_acc:
                                            em_file_acc.write(file_credit_card_em_acc+'|'+courier+'|Pri'+'\n')
                                        em_file_acc.close()
                                        
                                    elif card_holder_type_acc == '0':
                                            
                                        with open('xxAUF_FF_MIS_B'+str(c_bin_acc)+'_L'+str(logo_acc)+'_G'+str(gender_code_acc)+'_'+str(product).replace('AU_BANK_LIT_CREDIT_CARD','Lit_Credit_Card')+'_'+celeb+' - Add-On-Card.txt'+'.txt', 'a') as ff_addonfile_acc:
                                            ff_addonfile_acc.write(file_credit_card_ff_acc_addon+file_credit_card_ff_acc_2_addon+file_credit_card_ff_acc_3_addon+'|Add-On|'+exp_date_acc_ff+'|'+product_code+'\n')
                                        ff_addonfile_acc.close()
                                        
                                        with open('xxAUC_EM_B'+str(c_bin_acc)+'_L'+str(logo_acc)+'_G'+str(gender_code_acc)+'_'+str(product).replace('AU_BANK_LIT_CREDIT_CARD','Lit_Credit_Card')+'_'+celeb+'.txt', 'a') as em_file_acc:
                                            em_file_acc.write(file_credit_card_em_acc+'|'+courier+'|Add-On'+'\n')
                                        em_file_acc.close()
                
                                # =========================================================================================================================================================
                                if card_holder_type_acc == '1':
                                    with open('xxAUF_Credit_Data_Allocation_Report_.txt', 'a') as ff_file_acc:
                                        ff_file_acc.write(allocation_data_primary+'\n')
                                    ff_file_acc.close()
                                    
                                    with open('xxAUF_Credit_Scanning_File.txt', 'a') as ff_file_Scan:
                                        ff_file_Scan.write(scanning_data+'\n')
                                    ff_file_Scan.close()
                                    
                                    
                                elif card_holder_type_acc == '0':
                                    with open('xxAUF_Credit_Data_Allocation_Report_.txt', 'a') as ff_file_acc:
                                        ff_file_acc.write(allocation_data_addon+'\n')
                                    ff_file_acc.close()
                                    
                                    with open('xxAUF_Credit_Scanning_File.txt', 'a') as ff_file_Scan:
                                        ff_file_Scan.write(scanning_data+'\n')
                                    ff_file_Scan.close()
                                    
                                    
                                if logo_acc != '306':
                                    
                                    if ((c_bin_acc == '465523_1' and card_holder_type_acc=='1') or (c_bin_acc == '457036')  or (c_bin_acc == '653024' and logo_acc == '910') or (c_bin_acc == '653023' and  logo_acc == '912')):#and logo_acc == '402'):
                                    
                                    
                                        if (card_action_code == '1' or card_action_code == '2' or card_action_code == '7' or card_action_code == 'A'):
                                            if mask_pan_acc not in filecontents:
                                                with open('xxAUC_PP_B'+str(c_bin_acc)+'_L'+str(logo_acc)+'_G'+str(gender_code_acc)+'_'+card_action.replace(' ','-')+'_'+str(product).replace('AU_BANK_LIT_CREDIT_CARD','Lit_Credit_Card')+'_'+'EM.txt', 'a') as pp_file:
                                                    pp_file.write(file_pp+courier +'#7#'+card_action_1+ '\n')
                                                    #print (file_pp)
                                                pp_file.close()
                                                
                                                # not for znith plu as data is already present in primary ff
                                                if product != 'Zenith_Plus' or product != 'Zenith_Plus_RuPay' or product != 'AU_Traverse':
                                                    with open('xxAUC_PP_B'+str(c_bin_acc)+'_L'+str(logo_acc)+'_G'+str(gender_code_acc)+'_'+str(product).replace('AU_BANK_LIT_CREDIT_CARD','Lit_Credit_Card')+'_MIS.txt', 'a') as pp_file:
                                                        pp_file.write('|' +file_pp_merge+'|' +awb_number+'|'+courier+'|'+card_action_1+ '\n')
                                                        #print (file_pp)
                                                    pp_file.close()
                                            
                                                with open('AUC_PP_MERGE_'+str(ptime)+'.dat', 'a') as pp_file_m:
                                                    pp_file_m.write(file_pp_merge + '\n')
                                                    total_pp_count+=1
                                                pp_file_m.close()
                                            
                                                with open('Priority_Pass_Data_'+str(ptime[0:10])+'.csv', 'a') as pp_file_auf:
                                                    if count == 0:
                                                        pp_file_auf.write('')
                                                    pp_file_auf.write(auf_pp_output + '\n')
                                                pp_file_auf.close()
                                                    
                                
                                
                                            
    # =============================================================================
    #                                         with open('config/Priority_Pass_Data_AUF_BANK.csv', 'a') as pp_file_auf_ha:
    #                                             pp_file_auf_ha.write(auf_pp_output_bank+'\n')
    #                                         pp_file_auf_ha.close()
    # =============================================================================
                                            pp_list.append(auf_pp_output_bank)
                                
                                if (c_bin_acc == '483974' ):
                                    with open('ALERT!!___NEW BIN 483974 RECEIVED.dat', 'a') as alert_ff:
                                        alert_ff.write(c_bin_f_acc+'|'+logo_acc+'|'+gender_code_acc+'|'+varient+'|'+ref_number+'|'+card_type+'|'+mask_pan_acc+'|'+account_number_acc+'|'+card_action+'\n')
                                    alert_ff.close()
                                
                                
                                #Alert for new bin -483974
                                if (c_bin_acc == '483974' ):
                                    with open('ALERT!!___NEW BIN 483974 RECEIVED.dat', 'a') as alert_ff:
                                        alert_ff.write(c_bin_f_acc+'|'+logo_acc+'|'+gender_code_acc+'|'+varient+'|'+ref_number+'|'+card_type+'|'+mask_pan_acc+'|'+account_number_acc+'|'+card_action+'\n')
                                    alert_ff.close()
                                
                                if pct_id not in pct_id_val:
                                    with open('Wrong pct_id data_'+str(ptime)+'.txt','a') as data_file:
                                        if count == 0:
                                            data_file.write('Date,Account_Number,Logo_acc,Pct_id,Mask_pan_acc,Card_holder_name,card_action\n')
                                            count+=1
                                        data_file.write(pct_id_data + '\n')
                                        
                                    data_file.close()
                                            
                                            
                    
                    #print(account_number_list)
            
            outfile_acc.close()
            
            del_file2 = glob.glob("*account*")
            for del2 in del_file2:
                os.remove(del2)

            # 
            #         outfile.write()


            with open('xxAUF_Credit_Data_Allocation_Report_.txt', 'r') as file_alloc:
                contents = file_alloc.readlines()
                for line in contents:
                    awb = line.split('|')[13]
                    with open(f'AUF_CREDIT_SCANNING_CARD_FILE_{awb}.csv','a') as outfile:
                       outfile.write(line) 

            Primary_Card_count = ''
            files = glob.glob('AUF_CREDIT_SCANNING_CARD_FILE_*.csv')
            Srn = 0
            for file in files:
                
                addon_count =0
                previous_Primary_Card_number =''
                previous_Addon_Card_1 = ''
                previous_Addon_Card_2 = ''
                previous_Addon_Card_3 = ''
                previous_Addon_Card_4 = ''
                previous_PP_Card_No_Primary = ''
                previous_PP_Card_No_Addon1 = ''
                previous_PP_Card_No_Addon2 = ''
                previous_PP_Card_No_Addon3 = ''
                with open(file,'r') as infile,open('xAUF_CREDIT_SCANNING_CARD_FILE.txt','a') as outfile:
                    contents = infile.readlines()
                    size = len(contents)
                    for ln in contents:
                        val = ln.split('|')
                        Account_Number       =val[1].strip() 
                        Primary_Card_number  =val[2].strip()
                        Customer_Name        =val[3].strip()
                        Addon_Card_1         =val[4].strip()
                        Addon_Card_2          =val[5].strip()
                        Addon_Card_3         =val[6].strip()
                        Addon_Card_4         =val[7].strip()
                        PP_Card_No_Primary   =val[8].strip()
                        PP_Card_No_Addon1    =val[9].strip()  
                        PP_Card_No_Addon2    =val[10].strip()
                        PP_Card_No_Addon3    =val[11].strip()
                        Refno                =val[12].strip()
                        AWB_No               =val[13].strip()
                        Courier              =val[14].strip()
                        Courier_Code         =val[15].strip()
                        Card_Action          =val[16].strip()
                        Primary_Count        =val[17].strip()
                        Addon_Count          =val[18].strip()
                        Total_Cards          =val[19].strip()
                        Bin                  =val[20].strip()
                        Logo                 =val[21].strip()
                        Gender_Code          =val[22].strip()
                        Varient              =val[23].strip()
                        Address_Line_1       =val[24].strip()
                        Address_Line_2       =val[25].strip()
                        Address_Line_3       =val[26].strip()
                        Address_Line_4       =val[27].strip()
                        City                 =val[28].strip()
                        State_Code           =val[29].strip()
                        Postal_Code          =val[30].strip()
                        Mobile_Number        =val[31].strip()
                        Cust_Unique_ID       =val[32].strip()
                        Total_PP_Count       =val[33].strip()
                        Joining_Fees         =val[34].strip()
                        Annual_Monthly_Fees  =val[35].strip()
                        Card_Issuance_Date   =val[36].strip()
                        CARDHOLDER_EMBOSSA_2 =val[37].strip()
                        PLASTIC_ID           =val[38].strip()

                        if(Primary_Card_number != ''):
                            previous_Primary_Card_number = Primary_Card_number
                        if(Addon_Card_1 != ''):
                            previous_Addon_Card_1 = Addon_Card_1
                            addon_count +=1
                        if(Addon_Card_2 != ''):
                            previous_Addon_Card_2 = Addon_Card_2
                            addon_count +=1
                        if(Addon_Card_3 != ''):
                            previous_Addon_Card_3 = Addon_Card_3
                            addon_count +=1
                        if(Addon_Card_4 != '') :
                            previous_Addon_Card_4 = Addon_Card_4
                            addon_count +=1
                        if(PP_Card_No_Primary != ''):
                            previous_PP_Card_No_Primary = PP_Card_No_Primary
                        if(PP_Card_No_Addon1 != ''):
                            previous_PP_Card_No_Addon1 = PP_Card_No_Addon1
                        if(PP_Card_No_Addon2 != ''):
                            previous_PP_Card_No_Addon2 = PP_Card_No_Addon2
                        if(PP_Card_No_Addon2 != ''):
                            previous_PP_Card_No_Addon3 = PP_Card_No_Addon2
                        if previous_Primary_Card_number == '':
                            Primary_Card_count = '0'    
                        else:
                                Primary_Card_count = '1' 
                    Srn +=1
                    outfile.write(str(Srn).zfill(4)+'|'+Account_Number+'|'+previous_Primary_Card_number+'|'+previous_Addon_Card_1+'|'+previous_Addon_Card_2+'|'+previous_Addon_Card_3+'|'+previous_Addon_Card_4+'|'+previous_PP_Card_No_Primary+'|'+previous_PP_Card_No_Addon1+'|'+previous_PP_Card_No_Addon2+'|'+previous_PP_Card_No_Addon3+'|'+AWB_No+'|'+Courier+'|'+Primary_Card_count+'|'+str(addon_count)+'|'+str(size)+'|'+Bin+'|'+Logo+'|'+Gender_Code+'|'+Varient+'\n')
                            
                os.remove(file)
            




            print("\nEMBO and MIS files generated.....")
            print("\nSorting and excel conversion begins.....")
            for idx, val in enumerate(ps_list):
                #print(idx, val)
                acc_card_count = (ps_list.count(ps_list[idx]))
                if ps_list[idx][17:] == 'Primary Card':
                    p_acc_card_count.append(
                        str(ps_list[idx])+'|'+str(ps_list.count(ps_list[idx])))
                if ps_list[idx][17:] == 'Add-on  Card':
                    s_acc_card_count.append(
                        str(ps_list[idx])+'|'+str(ps_list.count(ps_list[idx])))
            
            with open('PrimaryCard_detail_'+str(ptime)+'.dat', 'w') as ps_d:
                ps_d.writelines(["%s\n" % item for item in p_acc_card_count])
            ps_d.close()
            
            
            with open('del_SecondaryCard.txt', 'w') as ps_d:
                ps_d.writelines(["%s\n" % item for item in s_acc_card_count])
            ps_d.close()
            
            uniqlines = set(open('del_SecondaryCard.txt').readlines())
            bar = open('SecondaryCard_details_'+str(ptime) +'.dat', 'w').writelines(set(uniqlines))
            
            for idx, val in enumerate(ps1_list):
                #print(idx, val)
                acc_card_count1 = (ps1_list.count(ps1_list[idx]))
                # if ps_list[idx][17:] == 'Primary Card':
                # p_acc_card_count.append(str(ps_list[idx])+'|'+str(ps_list.count(ps_list[idx])))
                if ps_list[idx][17:29] == 'Add-on  Card':
                    s_acc_card_count1.append(
                        str(ps1_list[idx])+'|'+str(ps1_list.count(ps1_list[idx])))
            
            
            with open('del_SecondaryCard1.txt', 'w') as ps_d:
                ps_d.writelines(["%s\n" % item for item in s_acc_card_count1])
                # for oni in s_acc_card_count1:
                #print ((oni))
            ps_d.close()
            
            uniqlines = set(open('del_SecondaryCard1.txt').readlines())
            bar = open('SecondaryCard1_details_'+str(ptime) +
                       '.txt', 'w').writelines(set(uniqlines))
            
            
            
            
            
            
            
            ff_filename = glob.glob("xxAUF_FF_MIS_B*")
            #print('File names:', ff_filename)
            for file in ff_filename:
                try:
                    excel_convertor(file)
                    log_write(file)
                except Exception as e:
                    print("Error while converting mis file : "+file+ "  "+ str(e))
            
            
            ff_filename = glob.glob("xxAUF_Replace_FF_MIS_B*")
            #print('File names:', ff_filename)
            for file in ff_filename:
                try: 
                    excel_convertor_replace(file)
                except Exception as e:
                    print("Error while converting replace mis file : "+file+ "  "+ str(e))
            
            
            
            
            
            embo_filename = glob.glob("xxAUC_EM_B*")
            #print('File names:', embo_filename)
            for file in embo_filename:
                try:
                    embo_convertor(file)
                except Exception as e:
                    print("Error while converting embo file : "+file+ "  "+ str(e))
            
            
            
            pp_embo_filename = glob.glob("xxAUC_PP_B*_EM.txt")
            #print('File names:', embo_filename)
            for file in pp_embo_filename:
                try: 
                    pp_embo_convertor(file)
                except Exception as e:
                    print("Error while converting pp embo file : "+file+ "  "+ str(e))
            
            
            pp_mis_filename = glob.glob("xxAUC_PP_B*_MIS.txt")
            #print('File names:', embo_filename)
            for file in pp_mis_filename:
                try:
                    pp_excel_convertor(file)
                except Exception as e:
                    print("Error while converting pp mis file : "+file+ "  "+ str(e))
            
                
            
           
            print("\nSorting and Excel Converted.....")
            print("\nCourier connection reports generating.....")
            IndiaPost_Courier = 'xxxAUF_Credit_IndiaPost_Courier_Connection_Report_.txt'
            if Path(IndiaPost_Courier).is_file():
                records_list = []
                s_no = 0
                with open('xxxAUF_Credit_IndiaPost_Courier_Connection_Report_.txt', "r") as ff, open('AUF_Credit_IndiaPost_Courier_Connection_Report_.txt', "w") as foutfile:
                    foutfile.write('SNO|Barcode_Value|REFERANCE NUMBER|CITY|PINCODE|NAME|ADDRESS 1|ADDRESS 2|ADDRESS 3|ADDRESSEE EMAIL|ADDRESSEE MOBILE|SENDER MOBILE|POD REQUIRED|Weight'+'\n')
                    contents = ff.readlines()
                    # sorting using our custom logic
                    # contents.sort(key=my_sort_ff)
                    for fi in contents:
                        records_list.append(fi)
                        s_no = (records_list.index(fi)+1)
                        s_no = "%05d" % s_no
                        if not fi.strip():
                            continue
                        if fi:
                            foutfile.write(str(s_no)+fi)
                ff.close()
                foutfile.close()
                df = pd.read_csv('AUF_Credit_IndiaPost_Courier_Connection_Report_.txt',encoding= 'unicode_escape', sep='|', dtype=object)
                df.to_excel('AUF_Credit_IndiaPost_Courier_Connection_Report_'+str(ptime[0:10])+'--' +str(s_no)+'.xlsx', 'Sheet1', index=False)
                os.remove('xxxAUF_Credit_IndiaPost_Courier_Connection_Report_.txt')
                os.remove('AUF_Credit_IndiaPost_Courier_Connection_Report_.txt')
            
            Delhivery_Courier = 'xxxAUF_Credit_Delhivery_Courier_Connection_Report_.txt'
            if Path(Delhivery_Courier).is_file():
                records_list=[]
                s_no=0
                with open('xxxAUF_Credit_Delhivery_Courier_Connection_Report_.txt',"r") as ff, open('AUF_Credit_Delhivery_Courier_Connection_Report_.txt',"w") as foutfile:
                    foutfile.write('Waybill|Order No/Reference No|Consignee Name|City|State|Country|Address|Pincode|Phone/Mobile|Weight|Payment Mode|Package Amount|Product to be Shipped|Return Address|Return Pin|Seller Name|Seller Address|person_specific|address_specific'+'\n')
                    contents = ff.readlines()
                    for fi in contents:
                        records_list.append(fi)
                        s_no= (records_list.index(fi)+1)
                        s_no = "%05d" % s_no
                        if not fi.strip():
                            continue
                        if fi:
                            foutfile.write(fi)
                ff.close()
                foutfile.close()
                
            
                df = pd.read_csv('AUF_Credit_Delhivery_Courier_Connection_Report_.txt',encoding= 'unicode_escape', sep='|', dtype=object)
                df.to_excel('AUF_Credit_Delhivery_Courier_Connection_Report_'+str(ptime[0:10])+'--'+str(s_no)+'.xlsx', 'Sheet1',index=False)
                os.remove('xxxAUF_Credit_Delhivery_Courier_Connection_Report_.txt')
                os.remove('AUF_Credit_Delhivery_Courier_Connection_Report_.txt')
            
            Bluedart_Courier = 'xxxAUF_Credit_Bluedart_Courier_Connection_Report_.txt'
            if Path(Bluedart_Courier).is_file():
                records_list = []
                s_no = 0
                with open('xxxAUF_Credit_Bluedart_Courier_Connection_Report_.txt', "r") as ff, open('AUF_Credit_Bluedart_Courier_Connection_Report_.txt', "w") as foutfile:
                    foutfile.write(
                        'Refrence Number|Customer Name|ADDRESS1|ADDRESS2|ADDRESS3|CITY|STATE|COUNTRY|PINCODE|CONTACT1|AWB'+'\n')
                    contents = ff.readlines()
                    for fi in contents:
                        records_list.append(fi)
                        s_no = (records_list.index(fi)+1)
                        s_no = "%05d" % s_no
                        if not fi.strip():
                            continue
                        if fi:
                            foutfile.write(fi)
                ff.close()
                foutfile.close()
                df = pd.read_csv('AUF_Credit_Bluedart_Courier_Connection_Report_.txt',encoding= 'unicode_escape', sep='|', dtype=object)
                df.to_excel('AUF_Credit_Bluedart_Courier_Connection_Report_' +str(ptime[0:10])+'--'+str(s_no)+'.xlsx', 'Sheet1', index=False)
                os.remove('xxxAUF_Credit_Bluedart_Courier_Connection_Report_.txt')
                os.remove('AUF_Credit_Bluedart_Courier_Connection_Report_.txt')
            
            
            folder_path = os.getcwd()
            password = "AU@321"

            # Iterate over all Excel files in the folder and protect them
            for filename in os.listdir(folder_path):
                if filename.__contains__('Courier_Connection_Report_') and filename.lower().endswith(".xlsx"):  # Only process .xlsx files
                    file_path = os.path.join(folder_path, filename)
                    # print(file_path)
                    protect_excel(file_path, password)
                    os.remove(filename)
            
            
            with open('xxAUF_FF_MIS_MERGE_'+str(ptime)+'.txt', "r") as ff, open('AUF_FF_MIS_MERGE_'+str(ptime)+'.txt', "w") as foutfile:
                records_list = []
                s_no = 0
                foutfile.write("Sr. No|FIRST SIX|Logo|Gender Code|Varient|Ref no.|CARD TYPE|MASK PAN|ACC NO|CARD ACTION|SALUTATION|MAILER NAME|Address Line 1|Address Line 2|Address Line 3|Address Line 4|City|State Code|Postal Code|Mobile Number|Cust Unique ID|Credit Limit|Statement Date |Extention|Embo_Name|PCT ID|AWB. No.|ROUTING CODE|Courier|Card Issuance Date"+'\n')
                contents = ff.readlines()
                # sorting using our custom logic
                # contents.sort(key=my_sort_ff)
                for fi in contents:
                    records_list.append(fi)
                    s_no = (records_list.index(fi)+1)
                    s_no = "%05d" % s_no
                    if not fi.strip():
                        continue
                    if fi:
                        foutfile.write(str(s_no)+fi)
                total_primary_count = s_no
            ff.close()
            foutfile.close()
            df = pd.read_csv('AUF_FF_MIS_MERGE_'+str(ptime)+'.txt',encoding= 'unicode_escape', sep='|', dtype=object)
            df.to_excel('AUF_FF_MIS_MERGE_'+str(ptime)+'-' +str(s_no)+'.xlsx', 'Sheet1', index=False)
            os.remove('xxAUF_FF_MIS_MERGE_'+str(ptime)+'.txt')
            os.remove('AUF_FF_MIS_MERGE_'+str(ptime)+'.txt')
            
            
            with open('xxAUF_Credit_Data_Allocation_Report_.txt', "r") as ff, open('AUF_Credit_Data_Allocation_Report_.txt', "w") as foutfile:
                records_list = []
                s_no = 0
                foutfile.write('Sr. No|Account Number|Primary Card number|Customer Name|Add-on Card 1|Add-on Card 2|Add-on Card 3|Add-on Card 4|PP  Card No_Primary|PP  Card No_Addon1|PP  Card No_Addon2|PP  Card No_Addon3|Ref no.|AWB. No.|Courier|Courier Code|Card Action|Primary Count|Add-on Count|Total Cards|Bin|Logo|Gender Code|Varient|Address Line 1|Address Line 2|Address Line 3|Address Line 4|City|State Code|Postal Code|Mobile Number|Cust Unique ID|Total PP Count|Joining Fees|Annual/Monthly Fees|Card Issuance Date|CARDHOLDER EMBOSSA_2|PLASTIC ID'+'\n')
                contents = ff.readlines()
                for fi in contents:
                    records_list.append(fi)
                    s_no = (records_list.index(fi)+1)
                    s_no = "%05d" % s_no
                    if not fi.strip():
                        continue
                    if fi:
                        foutfile.write(s_no+fi)
            ff.close()
            foutfile.close()
            df = pd.read_csv('AUF_Credit_Data_Allocation_Report_.txt',encoding= 'unicode_escape',sep='|', dtype=object)
            df.to_excel('AUF_Credit_Data_Allocation_Report_'+str(ptime[0:10]) +'--'+str(s_no) +'+'+str(total_pp_count)+'.xlsx', 'Sheet1', index=False)
            os.remove('xxAUF_Credit_Data_Allocation_Report_.txt')
            os.remove('AUF_Credit_Data_Allocation_Report_.txt')
            
            
            with open('xAUF_CREDIT_SCANNING_CARD_FILE.txt','r') as infile,open('AUF_CREDIT_SCANNING_CARD_FILE.csv','a') as scan_file:
                contents = infile.readlines()
                scan_file.write('Sr. No.|Account Number|Primary Card Number|Addon card 1|Addon card 2|Addon card 3|Addon card 4|PP Card No Primary|PP Card No Addon 1|PP Card No Addon 2|PP Card No Addon 3|Awb No.|Courier|Primary card Count|Addon Card Count|Total card Count|Bin|Logo|Gender Code|Varient'+'\n')
                scan_file.writelines(contents)       

            df = pd.read_csv('AUF_CREDIT_SCANNING_CARD_FILE.csv',encoding= 'unicode_escape', sep='|', dtype=object)
            df.to_excel('AUF_CREDIT_SCANNING_CARD_FILE_'+str(ptime[0:10]) +'--'+str(s_no) +'+'+str(total_pp_count)+'.xlsx', 'Sheet1', index=False)
            os.remove('AUF_CREDIT_SCANNING_CARD_FILE.csv')
            os.remove('xAUF_CREDIT_SCANNING_CARD_FILE.txt')
            
            
            
            courier_sum_total_count = str(s_no) +'+'+str(total_pp_count)
            
            files = glob.glob('AUF_Credit_Data_Allocation_Report*.xlsx')
            for fn in files:
                courier_summary(fn,str(ptime[0:10]),courier_sum_total_count)
            
            
            with open('xxAUF_Credit_Scanning_File.txt', "r") as ff, open('AUF_Credit_Scanning_File.txt', "w") as foutfile:
                records_list = []
                s_no = 0
                foutfile.write('Sr. No|Account Number|Primary Card number|Customer Name|Add-on Card 1|Add-on Card 2|Add-on Card 3|Add-on Card 4|PP  Card No_Primary|PP  Card No_Addon1|PP  Card No_Addon2|PP  Card No_Addon3|Ref no.|AWB. No.|Courier|Courier Code|Card Action|Primary Count|Add-on Count|Total Cards|Bin|Logo|Gender Code|Varient|Address Line 1|Address Line 2|Address Line 3|Address Line 4|City|State Code|Postal Code|Mobile Number|Cust Unique ID|Total PP Count|Joining Fees|Annual/Monthly Fees|Card Issuance Date|CARDHOLDER EMBOSSA_2|PLASTIC ID|Last 4 digit pan'+'\n')
                contents = ff.readlines()
                for fi in contents:
                    records_list.append(fi)
                    s_no = (records_list.index(fi)+1)
                    s_no = "%05d" % s_no
                    if not fi.strip():
                        continue
                    if fi:
                        foutfile.write(s_no+fi)
            ff.close()
            foutfile.close()
            df = pd.read_csv('AUF_Credit_Scanning_File.txt',encoding= 'unicode_escape',sep='|', dtype=object)
            df.to_excel('AUF_CREDIT_SCANNING_FILE_'+str(ptime[0:10]) +'--'+courier_sum_total_count+'.xlsx', 'Sheet1', index=False)
            os.remove('xxAUF_Credit_Scanning_File.txt')
            os.remove('AUF_Credit_Scanning_File.txt')
            
            
            del_file = glob.glob("merge*")
            for dem in del_file:
                os.remove(dem)
            del_file1 = glob.glob("del*")
            for de in del_file1:
                os.remove(de)
            
            
            print("\nCourier connection reports generated.....")
            print("\nBatch Card generating.....")
            
            
            try:
                with open('config/AUC_batch_series', 'r') as fin:
                    batch_series_data = fin.readlines()
                    batch_number = batch_series_data[0].rstrip()
                fin.close()
            except:
                print("Batch Series Finished!!!!")
            
            
            
            for file in glob.glob('*.txtt'):
                oldName = file
                newName = file.replace(".txtt",f"{batch_number}_.txtt")
                os.rename(oldName, newName)
            # cardnumber_acc = line_acc[9:25]
            for f in glob.glob('*.txtt'):
                with open(f,'r') as infile,open('xAUF_CREDIT_FPA.txt','a') as out_fpa:
                    contents = infile.readlines()
                    for line in contents:
                        # val = line.split('^')
                        # branch_code = val[-4].strip()
                        card_no =  line[19:25]
                        out_fpa.write(card_no+'|'+f+'|'+str(datedmyd)+'\n')
                
            with open('xAUF_CREDIT_FPA.txt','r') as fpa_in, open('AUF_CREDIT_FPA.csv','a') as fpa_out:
                contents = fpa_in.readlines()
                size = len(contents)
                sn = 0
                fpa_out.write('Sr. No.|LAST SIX CARD NO.|FILENAME|Date\n')
                for line in contents:
                    sn += 1
                    fpa_out.write(str(sn).zfill(4)+'|'+line)
            

            df = pd.read_csv('AUF_CREDIT_FPA.csv',sep='|', dtype=object, encoding="windows-1252")
            df.to_excel('AUF_CREDIT_FPA_SCANNING_FILE_'+batch_number+'_'+str(datedmyd)+'_'+str(size).zfill(4)+'.xlsx', 'Sheet1', index=False)
            os.remove('xAUF_CREDIT_FPA.txt')
            os.remove('AUF_CREDIT_FPA.csv')

            import shutil
            with open('config\FPA_FILE_LOCATION.csv','r') as fpa_loc:
                content = fpa_loc.readlines()
            destination_folder = content[0]
            os.makedirs(destination_folder, exist_ok=True)
            files = glob.glob('AUF_CREDIT_FPA_SCANNING_FILE_*.xlsx') 
            for file in files:
                destination_path = os.path.join(destination_folder, os.path.basename(file))
                shutil.move(file, destination_path)


            for i,k in enumerate(file_name):
                result = hashlib.md5(file_name[i].encode())
                filenamehash = result.hexdigest()
                filenamehash = str(filenamehash)
                #print (file_name[i]+'---'+filenamehash)
                with open('C:/AUTO-PROCESS-CONFIG/AUC/AUC_PROCESSING_'+str(ptime)+'.log', 'a') as hashfile:
                    hashfile.write("File Processed  ")
                    hashfile.write(file_name[i]+'|'+str(ptime)+'|'+filenamehash+'  ')
                    hashfile.write("File Deleted \n")
                hashfile.close()
                with open('C:/AUTO-PROCESS-CONFIG/AUC/AUC_PROCESSING.log', 'a') as hashfilem:
                    hashfilem.write("File Processed  ")
                    hashfilem.write(file_name[i]+'|'+str(ptime)+'|'+filenamehash+'  ')
                    hashfilem.write("File Deleted \n")
                hashfilem.close()
                     
            file_count = '.'
            
            
            os.chdir(file_count)
            names={}
            for fn in glob.glob('*AUC*.txtt'):
                with open(fn) as f:
                    names[fn]=sum(1 for file_count in f if file_count.strip() and not file_count.startswith('~'))       
            with open('AUF_FILE_COUNT_'+str(ptime)+'.csv', 'w') as f:
                f.write('File Name,File Count,Non-Replace,Replace'+'\n')
                [f.write('{0},{1}\n'.format(key, value)) for key, value in names.items()] 
            
            from prettytable import PrettyTable
            ptx = PrettyTable()
            ptx.field_names = ["Bin", "Artwork No.","JB No.", "EMBOSS Filename","Qty", "Job setup", "Printing method - Front/Back"]
            
            for key, value in names.items():
                art_work = 'Not_Found'
                job_setup = 'Not_Found'
                dg_color = 'Not_Found'
                with open('config/batchcard.txt','r') as data_file, open('config/batchcard_file.csv','r') as batch_file:
                    contents = data_file.readlines() 
                    datas = batch_file.readlines() 
                    
                    for line in contents[1:]:
                        line_data = line.split(',')
                        file_start_index = line_data[0].strip()
                        file_end_index = line_data[1].strip()
                        file_embo_name = line_data[2].strip()
                        file_art_work = line_data[3].strip()
                        file_job_setup = line_data[4].strip()
                        file_dg_color = line_data[5].strip()
                        
                        # print(file_start_index+':'+file_end_index)
                        if key[int(file_start_index):int(file_end_index)] == file_embo_name:
                            art_work = file_art_work
                            job_setup = file_job_setup
                            dg_color = file_dg_color
                            break
                        
                    for data in datas[1:]:
                        value_1 = data.split(',')
                        bin_num = value_1[0].rstrip()
                        job_setup_batch = value_1[1].rstrip()
                        file_name_batch = value_1[2].rstrip()
                        logo_batch = value_1[3].rstrip()
                        art_work_batch = value_1[4].rstrip()
                        
                        if (bin_num == key[8:14] and job_setup_batch == job_setup and  art_work_batch == art_work or file_name_batch in key):
                            logo = key.split('_')[3][1:]
                            card_action_e = key.split('_')[5]
                            # print(logo)
                            printing_front = value_1[5].rstrip()
                            ribbon_front = value_1[6].rstrip()
                            printing_back = value_1[7].rstrip()
                            ribbon_back = value_1[8].rstrip()
                            
                            with open('AUF_CREDIT_BATCHCARD_FILE.csv','a') as out_file:
                                out_file.write(''+'|'+str(value)+'|'+key+'|'+job_setup_batch+'|'+art_work_batch+'|'+printing_front+'|'+ribbon_front+'|'+printing_back+'|'+ribbon_back+'|'+bin_num+'|'+logo+'\n')
                            break
                ptx.add_row([key[8:14], art_work,"", key,value, job_setup, dg_color])
                
            
            ptx.align = "c"
            ptd=ptx.get_string()
            
            ptx1 = PrettyTable()
            ptx1.field_names = ["Emboss Filename","Supervisor Name","Signature", "  Date  ","  Time  ","   Remark(if any)   "]
            for key, value in names.items():
                
                ptx1.add_row([key,"","","","",""])
            ptx1.align = "c"
            ptd1=ptx1.get_string()
            qty1 = sum(names.values())
            datedmy = str(x)[8:10]+'-'+str(x)[5:7]+'-'+str(x)[0:4]
            with open('AUF_CREDIT_BATCHCARD_'+str(ptime)+'.dat', 'w') as file:
              file.write('                                                '+'BANKING PERSONALISATION BATCH CARD'+'                              '+'Date: '+ptime[0:10]+'\n')
              file.write('                                                       AUC-'+ptime[0:10]+'-'+batch_number+'                                    '+'AUF CREDIT PROJECT(DI)\n\n')
              file.write(str(ptd))
              file.write('\n')
              file.write('TOTAL BATCH QUANTITY:'+str(sum(names.values()))+'\n')
              file.write('                                                   Data upload on Machine\n')
              file.write(str(ptd1))
              file.write('\n\n')
              file.write('PRP: 20.1                                              Rev No: 3.3                                            Date: 01-FEB-25\n')
              file.write('SEC-3: INTERNAL                                        Owner: Quality Control                                 Status: Issued\n\n')
              file.write('                                                       Page: 1 of 1')
                 
            
            
            from fpdf import FPDF
            class PDF(FPDF):
                def header(self):
                    self.image('config/colorplast_logo.png', x = 10, y = 5, w = 50, h = 10, type = '', link = '')
                    self.ln(8)
            pdf = PDF()
            pdf.add_page('L')
            pdf.set_font("Courier", size=6.5)
            f = open('AUF_CREDIT_BATCHCARD_'+str(ptime)+'.dat', "r")
            
            for x in f:
                pdf.cell(50, 5, txt = x, ln=True, align = 'L')
                #pdf.image('config/colorplast_logo.png', x = 10, y = 5, w = 50, h = 20, type = '', link = '')
                #pdf.cell(ln=10)
            
            pdf.output('AUF_CREDIT_BATCHCARD_'+ptime[0:10]+'_'+batch_number+'.pdf')
            f.close()
            
            with open('AUF_CREDIT_BATCHCARD_FILE.csv','r') as in_file, open('AUF_CREDIT_BATCHCARD_'+ptime[0:10]+'_'+str(sum(names.values()))+'.TXT','a') as output_file:
                contents = in_file.readlines()
                output_file.write('CARD_ACTION|QTY|Emboss_Filename|JOB_SETUP_NAME|ARTWORK NO.|F-PRINTING_Method|F-Ribbon_Color|B-PRINTING_Method|B-Ribbon_Color|BIN|SUB_BIN/LOGO\n')
                sn = 0
                for line in contents:
                    sn+=1
                    output_file.write(str(sn)+'|'+line)
                    
            df = pd.read_csv('AUF_CREDIT_BATCHCARD_'+ptime[0:10]+'_'+str(sum(names.values()))+'.TXT',encoding= 'unicode_escape',sep='|', dtype=object)
            df.to_excel('AUF_CREDIT_BATCHCARD_'+datedmyd+'_'+str(sum(names.values()))+'.xlsx', 'Sheet1', index=False)       
            

            input_file = glob.glob('AUF_CREDIT_BATCHCARD_*.xlsx')
            process_excel_file(input_file[0])

            #---------------------------------BatchCard ENDS ---------------------------
            
            #---------------------------------DELETION LOG-----15.2.2----------------------
            from prettytable import PrettyTable
            ptx = PrettyTable()
            ptx.field_names = (["Client Name", "File Name","File Deletion Date", "Dispatch Date","Data Admin. Name", "Data Admin. Sign.","IT Person Name", "IT Person Sign."])
            
            for key,value in names.items():
                
                ptx.add_row(["AUF Bank", key, "", "", "Data Team", "", "IT Team", ""])
            ptx.align = "l"
            ptd=ptx.get_string()
            # print(ptd)
            with open('AUF_DELETION_LOG_'+datedmy+'.dat', 'w') as file:
              file.write('\n\n                                                                        SEC-IS   |   02.02\n')
              file.write('                                                                        Rev.No.  |   3.0                                                   '+'AUF CREDIT PROJECT(DI)\n')
              file.write('                                                                        Date :   |   15-Jul-18\n')
              file.write('                                                                 '+'( ***** DATA DELETION LOG ***** )''\n\n')
              file.write('|  Data Receiving Date - '+ptime[0:10]+'  |  '+'  Data Server - DPP Server/MX Machine  |   '+'  Batch Qty. - '+str(sum(names.values()))+'   |  '+'  Batch No. - AUC-'+ptime[0:10]+'-'+batch_number+'    |  '+'  Status -                  '+'|''\n')
              
              file.write(str(ptd).rstrip())
              file.write('\n')
              #file.write('BATCH NO. OF PM TOOL FOR THIS BATCHCARD :_______________\n')
            # from fpdf import FPDF
            class PDF1(FPDF):
                def header(self):
                    self.image('config/colorplast_logo.png', x = 10, y = 5, w = 50, h = 10, type = '', link = '')
                    self.ln(8)
            
            pdf1 = PDF1()
            pdf1.add_page('L')
            pdf1.set_font("Courier", size=6.5)
            f = open('AUF_DELETION_LOG_'+datedmy+'.dat', "r")
            
            for x in f:
                pdf1.cell(50, 5, txt = x, ln=True, align = 'L')
                #pdf.image('config/colorplast_logo.png', x = 10, y = 5, w = 50, h = 20, type = '', link = '')
                #pdf.cell(ln=10)
            
            
            pdf1.output('AUF_DELETION_LOG_'+ptime[0:10]+'.pdf')
            f.close()
            
            #---------------------------------DELETION LOG ENDS ---------------------------
        
            print("Batch Generated..........")
            
            import shutil
            destination_folder = "C:/Config/Batchcard"
            os.makedirs(destination_folder, exist_ok=True)
            files = glob.glob('*batchcard*.dat') +glob.glob('*deletion*.dat')
            for file in files:
                destination_path = os.path.join(destination_folder, os.path.basename(file))
                shutil.move(file, destination_path)
            
            path = ('.')
            ext = "txtt"
            
            for f in os.listdir(path):
                fpath = os.path.join(path, f)
            
                if os.path.isfile(fpath) and fpath.endswith(ext):
                    # ctime = datetime.datetime.fromtimestamp(os.path.getctime(fpath)).strftime(ptime_all+str(s_no) +'+'+str(total_pp_count))
                    name='AUC_'+ptime_all+'---'+courier_sum_total_count
                    os.makedirs(os.path.join(path, name), exist_ok=True)
                    os.makedirs(os.path.join("A:/Sdrive/LIVE/AU_Credit/", name), exist_ok=True)
                    os.replace(fpath, os.path.join(path, name, f))
            
            
            
               
            from distutils.dir_util import copy_tree        
            
            fromDirectory =  os.getcwd()
            aaa = fromDirectory+'/'+name
            
            toDirectory = "A:/Sdrive/LIVE/AU_Credit/"
            bbb = toDirectory +name
            # print(aaa)
            # print(bbb)
            copy_tree(aaa, bbb)
            
            print ('\n\n.......Files Processed......Total Credit Card Records: '+str(len(records_list)))
            
            
            for file in glob.glob('*.xlsx'):
                oldName = file
                newName = file.replace("_-","-")
                os.rename(oldName, newName)
            
            del_file = glob.glob("L2CDZU*")
            for dem in del_file:
                os.remove(dem)
            
            with open("config/File_DupeCheck.csv",'r+') as dupecheck_file:
                date_now=int(datetime.datetime.now().strftime("%y%j"))
                dupecheck_file_filecontents = dupecheck_file.readlines()
                dupecheck_file.seek(0)
                for line in dupecheck_file_filecontents:
                    date_check = int(line.split(":")[0].rstrip())
                    
                    if (date_now-15)<date_check:
                        
                        dupecheck_file.write(line.rstrip()+"\n")
                for file in file_name:
                    print(file)
                    dupecheck_file.write(str(datetime.datetime.now().strftime("%y%j"))+":"+file+"\n")
                
                dupecheck_file.truncate()
            dupecheck_file.close()
                
            with open('config/BD_AWB.txt', 'r+') as bdfin:
                contents =  bdfin.readlines()
                bdremainingnew = len(contents[bd_num:])
                bdfin.seek(0)
                for line in contents[bd_num:]:
                    bdfin.write(line)
                bdfin.truncate()
            bdfin.close()
            
            with open('config/DL_AWB.txt', 'r+') as dlfin:
                contents =  dlfin.readlines()
                dtremainingnew = len(contents[dl_num:])
                dlfin.seek(0)
                for line in contents[dl_num:]:
                    dlfin.write(line)
                dlfin.truncate()
            dlfin.close()
            
           
            with open('config/IP_AWB.txt', 'r+') as ipfin:
                contents =  ipfin.readlines()
                ipremainingnew = len(contents[ip_num:])
                ipfin.seek(0)
                for line in contents[ip_num:]:
                    ipfin.write(line)
                ipfin.truncate()
            ipfin.close()
                
            with open('AWB_REMAINING_COUNT_NEW.txt','w') as awb_count_new:
                awb_count_new.write('BD   : ' + str(bdremainingnew)+'\n'+'DL : '+str(dtremainingnew)+'\nIP   : '+str(ipremainingnew) )
                print('BD   : ' + str(bdremainingnew)+'\n'+'DL : '+str(dtremainingnew)+'\nIP   : '+str(ipremainingnew) )
            awb_count_new.close()
            
            
            
            with open('config/VETTA_PP_NUMBER.csv', 'r+') as vetta_pp_in:
                contents =  vetta_pp_in.readlines()
                vetta_remainingnew = len(contents[vetta_pp:])
                vetta_pp_in.seek(0)
                for line in contents[vetta_pp:]:
                    vetta_pp_in.write(line)
                vetta_pp_in.truncate()
            vetta_pp_in.close()
        
            
            with open('config/ZENITH_PP_NUMBER.csv', 'r+') as zenith_pp_in:
                contents = zenith_pp_in.readlines()
                zenith_remainingnew = len(contents[zenith_pp:])
                zenith_pp_in.seek(0)
                for line in contents[zenith_pp:]:
                    zenith_pp_in.write(line)
                zenith_pp_in.truncate()
                
            zenith_pp_in.close()
            
            with open('config/AUC_batch_series', 'r+') as batch_series_read:
                contents =  batch_series_read.readlines()
                batch_series_read.seek(0)
                for line in contents[1:]:
                    batch_series_read.write(line)
                batch_series_read.truncate()
            batch_series_read.close()
            
            current_directory = os.getcwd()
            
            
            
            with open('REMAINING_PP_COUNT_NEW.txt','w') as pp_count_new:
                pp_count_new.write('Remaining VETTA Count  : ' + str(vetta_remainingnew)+'\n'+'Remaining ZENITH Count : '+str(zenith_remainingnew)+'\n' )
                print('Remaining VETTA  Count  : ' + str(vetta_remainingnew)+'\n'+'Remaining ZENITH Count : '+str(zenith_remainingnew)+'\n')
            pp_count_new.close()
            
            with open('./config/AUF_CREDIT_PP_CONSOLE.txt','a') as file_write:
                for line in pp_num:
                    file_write.write(line+'\n')
                    
                    
            file_write.close()
            
            with open('config/Priority_Pass_Data_AUF_BANK.csv', 'a') as pp_file_auf_ha:
                for data in pp_list:
                    pp_file_auf_ha.write(data+"\n")
            pp_file_auf_ha.close()

            del_file = glob.glob("xxAUF*") 
            for dem in del_file:
                os.remove(dem)
            
            
            del_file = glob.glob("*batchcard*.csv") + glob.glob("*batchcard*.txt")
            for dem in del_file:
                os.remove(dem)
            
            excel_files = glob.glob(os.path.join(current_directory, '*FF*.xlsx')) + glob.glob(os.path.join(current_directory, '*PP*.xlsx')) + glob.glob(os.path.join(current_directory, '*Scan*.xlsx'))+glob.glob(os.path.join(current_directory, '*Courier_Connection*.xlsx'))

            if excel_files:
                for excel_file in excel_files:
                    convert_to_read_only(excel_file)
                    
            with open('config/Output_file_location.csv','r') as file:
                content = file.readlines()
                folder_location = content[0].split(',')[1].strip()
                pp_location = content[1].split(',')[1].strip()
                Customer_location = content[2].split(',')[1].strip()
                path = os.getcwd()
                files = glob.glob('Priority_Pass_Data*.csv') + glob.glob('Special_Character_Records*.csv') + glob.glob('AUF_Credit_Data_Allocation_Report*.xlsx') + glob.glob('AUF_File_DupeCheck_*.csv') + glob.glob('AUC_PROCESSING.log')
                for f in files:
                    if f.__contains__('Priority_Pass_Data'):
                        os.makedirs(pp_location, exist_ok=True)
                        shutil.copy(f, pp_location)
                    else:
                        
                        os.makedirs(Customer_location, exist_ok=True)
                        shutil.copy(f, Customer_location)         
                        
                        
        
            path = ('.')
           
            for f in os.listdir(path):
                fpath = os.path.join(path, f)
                if os.path.isfile(fpath) and (fpath.endswith(('xlsx', 'pdf', 'dat', 'csv', 'txt', 'log','xls'))):              
                    name='AUF_CREDIT_FF_MIS_'+ptime_all+'--'+courier_sum_total_count
                    os.makedirs(os.path.join(path, name), exist_ok=True)
                    file_name, file_extension = os.path.splitext(fpath)
                    os.replace(f'{fpath}', f'{file_name}_{batch_number}{file_extension}')
                    fpath =f'{file_name}_{batch_number}{file_extension}' 
                    os.replace(fpath, os.path.join(path, name.replace('.txt','.txtttt'), f))
                   
                      
            os.chdir(path) 
            fromDirectory = os.getcwd()
            aaa = fromDirectory+'/'+name
            toDirectory = folder_location 
            bbb = toDirectory+name
            copy_tree(aaa, bbb)          
                    
except Exception as e:
    traceback.print_exc()
    
       
    
ts = datetime.datetime.now()
ptime = ts.strftime("%d.%m.%Y_%H%M%S")
print(ptime)

print("Closing...")    
time_limit = 10800  # 3 hours in seconds
start_time = time.time()  # This should be fine since we've imported time module

while True:
    if time.time() - start_time >= time_limit:
        print("Time limit reached. Exiting program.")
        break
    time.sleep(1)